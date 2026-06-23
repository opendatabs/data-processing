import io
import logging
import math
import os
import re
import zipfile
from dataclasses import dataclass
from datetime import UTC, date, datetime
from pathlib import Path
from typing import Any
from zoneinfo import ZoneInfo

import common
import pandas as pd
import requests
from dotenv import load_dotenv
from openpyxl import load_workbook
from openpyxl.styles import Font

load_dotenv()

LOGGER = logging.getLogger(__name__)

DASHBOARD_BASEL = "Basel"
TARGET_ATTRIBUTE_NAMES = frozenset({"Ecoli", "Ecoli_2", "Entero", "Transmission", "Gehalt"})
TARGET_ATTRIBUTE_NAMES_LOWER = {n.lower() for n in TARGET_ATTRIBUTE_NAMES}
# Stable column order for the wide Coliminder frame and the persisted Excel file.
COLIMINDER_VALUE_COLUMNS = ["Ecoli", "Ecoli_2", "Entero", "Transmission", "Gehalt"]
EXPORT_FROM_DATE = date(2026, 4, 28)
COLIMINDER_EXPORT_FROM = date(2026, 1, 1)

DATA_DIR = Path("data")
OGD_DIR = DATA_DIR / "daten_aus_ogd"
LAB_EXCEL = DATA_DIR / "input" / "Laborergebnisse2026.xlsx"
COLIMINDER_XLSX = DATA_DIR / "etl_output" / "Coliminder.xlsx"
JOINED_XLSX = DATA_DIR / "etl_output" / "joined_data.xlsx"
LEGACY_SHORT_XLSX = DATA_DIR / "input" / "data_ecoli_entero_short.xlsx"
TZ = ZoneInfo("Europe/Zurich")


@dataclass(frozen=True)
class Config:
    base_url: str
    realm: str
    client_id: str
    client_secret: str

    @classmethod
    def from_env(cls) -> "Config":
        required_vars = {
            "API_COLIMINDER_BASE_URL": os.getenv("API_COLIMINDER_BASE_URL"),
            "API_COLIMINDER_REALM": os.getenv("API_COLIMINDER_REALM"),
            "API_COLIMINDER_CLIENT_ID": os.getenv("API_COLIMINDER_CLIENT_ID"),
            "API_COLIMINDER_CLIENT_SECRET": os.getenv("API_COLIMINDER_CLIENT_SECRET"),
        }
        missing = [name for name, value in required_vars.items() if not value]
        if missing:
            missing_text = ", ".join(missing)
            raise ValueError(f"Missing required environment variable(s): {missing_text}")

        return cls(
            base_url=required_vars["API_COLIMINDER_BASE_URL"].rstrip("/"),
            realm=required_vars["API_COLIMINDER_REALM"],
            client_id=required_vars["API_COLIMINDER_CLIENT_ID"],
            client_secret=required_vars["API_COLIMINDER_CLIENT_SECRET"],
        )


class OpenRemoteClient:
    """Read-only OpenRemote API client (no attribute writes)."""

    def __init__(self, config: Config, timeout_seconds: int = 30) -> None:
        self.config = config
        self.timeout_seconds = timeout_seconds
        self.session = requests.Session()
        self.token_endpoint_used: str | None = None
        self.api_base_path_used: str | None = None

    def authenticate(self) -> None:
        token_candidates = (
            f"{self.config.base_url}/auth/realms/{self.config.realm}/protocol/openid-connect/token",
            f"{self.config.base_url}/realms/{self.config.realm}/protocol/openid-connect/token",
        )
        payload = {
            "grant_type": "client_credentials",
            "client_id": self.config.client_id,
            "client_secret": self.config.client_secret,
        }
        headers = {"Content-Type": "application/x-www-form-urlencoded"}

        errors: list[str] = []
        for token_url in token_candidates:
            try:
                response = self.session.post(
                    token_url,
                    data=payload,
                    headers=headers,
                    timeout=self.timeout_seconds,
                )
            except requests.RequestException as exc:
                errors.append(f"{token_url} -> request error: {exc}")
                continue

            if response.ok:
                body = response.json()
                token = body.get("access_token")
                if not token:
                    errors.append(f"{token_url} -> no access_token in response")
                    continue
                self.session.headers.update({"Authorization": f"Bearer {token}"})
                self.token_endpoint_used = token_url
                LOGGER.info("Authentication succeeded with token endpoint: %s", token_url)
                return

            errors.append(f"{token_url} -> HTTP {response.status_code}: {response.text[:300]}")

        error_text = " | ".join(errors)
        raise RuntimeError(f"Authentication failed for all token endpoints. Details: {error_text}")

    def _request(
        self,
        method: str,
        endpoint: str,
        *,
        json_body: Any = None,
        params: dict[str, Any] | None = None,
    ) -> Any:
        api_base_candidates = (
            f"{self.config.base_url}/api/master",
            f"{self.config.base_url}/api/manager/v1",
        )
        errors: list[str] = []
        for api_base in api_base_candidates:
            url = f"{api_base}{endpoint}"
            try:
                response = self.session.request(
                    method=method,
                    url=url,
                    json=json_body,
                    params=params,
                    timeout=self.timeout_seconds,
                )
            except requests.RequestException as exc:
                errors.append(f"{url} -> request error: {exc}")
                continue

            if response.ok:
                self.api_base_path_used = api_base
                content_type = response.headers.get("content-type", "")
                if "application/json" in content_type:
                    return response.json()
                return response.text

            if response.status_code in (401, 403):
                errors.append(f"{url} -> HTTP {response.status_code} (not authorized)")
                continue

            if response.status_code == 404:
                errors.append(f"{url} -> HTTP 404 (endpoint not found)")
                continue

            errors.append(f"{url} -> HTTP {response.status_code}: {response.text[:300]}")

        error_text = " | ".join(errors)
        raise RuntimeError(f"Request failed for endpoint {endpoint}. Details: {error_text}")

    def get_dashboards(self) -> Any:
        return self._request("GET", f"/dashboard/all/{self.config.realm}")

    def query_assets(self) -> Any:
        return self._request("POST", "/asset/query", json_body={})

    def get_datapoints(
        self,
        asset_id: str,
        attribute_name: str,
        from_timestamp_ms: int,
        to_timestamp_ms: int,
    ) -> Any:
        endpoint = f"/asset/datapoint/{asset_id}/{attribute_name}"
        body = {
            "fromTimestamp": from_timestamp_ms,
            "toTimestamp": to_timestamp_ms,
        }
        return self._request("POST", endpoint, json_body=body)


def _extract_items(payload: Any) -> list[dict[str, Any]]:
    if isinstance(payload, list):
        return [item for item in payload if isinstance(item, dict)]
    if isinstance(payload, dict):
        for key in ("data", "items", "results"):
            maybe_items = payload.get(key)
            if isinstance(maybe_items, list):
                return [item for item in maybe_items if isinstance(item, dict)]
    return []


def _find_dashboard_by_display_name(payload: Any, display_name: str) -> dict[str, Any] | None:
    target = display_name.strip().casefold()
    for item in _extract_items(payload):
        name = item.get("displayName")
        if isinstance(name, str) and name.strip().casefold() == target:
            return item
    return None


def _walk_collect_attribute_refs(obj: Any, acc: list[tuple[str, str]]) -> None:
    """Collect (asset_id, attribute_name) from OpenRemote AttributeRef lists (id + name)."""
    if isinstance(obj, dict):
        refs = obj.get("attributeRefs")
        if isinstance(refs, list):
            for ref in refs:
                if not isinstance(ref, dict):
                    continue
                asset_id = ref.get("id")
                attr_name = ref.get("name")
                if isinstance(asset_id, str) and isinstance(attr_name, str):
                    acc.append((asset_id, attr_name))
        for value in obj.values():
            _walk_collect_attribute_refs(value, acc)
    elif isinstance(obj, list):
        for element in obj:
            _walk_collect_attribute_refs(element, acc)


def _basel_dashboard_attribute_pairs(dashboard: dict[str, Any]) -> list[dict[str, str]]:
    raw: list[tuple[str, str]] = []
    _walk_collect_attribute_refs(dashboard, raw)
    seen: set[tuple[str, str]] = set()
    pairs: list[dict[str, str]] = []
    for asset_id, attribute_name in raw:
        if attribute_name.lower() not in TARGET_ATTRIBUTE_NAMES_LOWER:
            continue
        key = (asset_id, attribute_name)
        if key in seen:
            continue
        seen.add(key)
        pairs.append({"asset_id": asset_id, "attribute_name": attribute_name})
    return pairs


def _build_asset_id_to_name(assets_payload: Any) -> dict[str, str]:
    mapping: dict[str, str] = {}
    for asset in _extract_items(assets_payload):
        asset_id = asset.get("id")
        if not asset_id:
            continue
        mapping[str(asset_id)] = str(asset.get("name", ""))
    return mapping


def _normalize_datapoints(payload: Any) -> list[dict[str, Any]]:
    if isinstance(payload, list):
        return [item for item in payload if isinstance(item, dict)]
    if isinstance(payload, dict):
        for key in ("data", "items", "results", "datapoints"):
            points = payload.get(key)
            if isinstance(points, list):
                return [item for item in points if isinstance(item, dict)]
    return []


def _find_column(columns: list[Any], *needles: str) -> str | None:
    """Return first column name matching all needles (case-insensitive, ascii-folded)."""

    def fold(s: str) -> str:
        s = str(s).lower()
        s = s.replace("ü", "u").replace("ö", "o").replace("ä", "a").replace("ß", "ss")
        return re.sub(r"[^a-z0-9]+", "", s)

    folded_needles = [fold(n) for n in needles]
    for col in columns:
        f = fold(col)
        if all(n in f for n in folded_needles):
            return str(col)
    return None


def load_labor_excel(path: Path | str = LAB_EXCEL) -> pd.DataFrame:
    """Load Kantonslabor Excel; build tz-aware ``Messzeitpunkt_Labor`` (Europe/Zurich)."""
    path = Path(path)
    df = pd.read_excel(path)
    df.columns = [str(c).strip() for c in df.columns]

    col_datum = _find_column(list(df.columns), "datum") or "Datum"
    col_zeit = _find_column(list(df.columns), "uhrzeit") or "Uhrzeit"
    col_ecoli = _find_column(list(df.columns), "e", "coli") or "E.coli"
    col_entero = _find_column(list(df.columns), "entero") or "Enterokokken"
    col_bwq = _find_column(list(df.columns), "badewasser") or "Badewasserqualität"
    col_ort = _find_column(list(df.columns), "entnahme") or "Entnahmeort"

    datum_series = pd.to_datetime(df[col_datum], errors="coerce")
    ut = df[col_zeit]
    if pd.api.types.is_timedelta64_dtype(ut):
        ut_str = (pd.Timestamp("1970-01-01", tz=TZ) + ut).dt.strftime("%H:%M:%S")
    elif pd.api.types.is_datetime64_any_dtype(ut):
        ut_str = pd.to_datetime(ut).dt.strftime("%H:%M:%S")
    else:
        ut_str = ut.astype(str).str.strip()

    combined = pd.to_datetime(
        datum_series.dt.strftime("%Y-%m-%d") + " " + ut_str,
        errors="coerce",
    )
    mess = combined.dt.tz_localize(TZ, ambiguous="infer", nonexistent="shift_forward")

    out = pd.DataFrame(
        {
            "Messzeitpunkt_Labor": mess,
            "Entnahmeort": df[col_ort] if col_ort in df.columns else pd.NA,
            "E.coli": df[col_ecoli] if col_ecoli in df.columns else pd.NA,
            "Enterokokken": df[col_entero] if col_entero in df.columns else pd.NA,
            "Badewasserqualität": df[col_bwq] if col_bwq in df.columns else pd.NA,
        }
    )
    return out


def load_legacy_short_excel(path: Path | str = LEGACY_SHORT_XLSX) -> pd.DataFrame:
    """Load old `data_ecoli_entero_short.xlsx` if available."""
    p = Path(path)
    if not p.exists():
        return pd.DataFrame(
            columns=[
                "Messzeitpunkt_Labor",
                "Entnahmeort",
                "E.coli",
                "Enterokokken",
                "Badewasserqualität",
                "coliminder_e_coli",
                "coliminder_entero",
            ]
        )

    df = pd.read_excel(p)
    df.columns = [str(c).strip() for c in df.columns]

    col_datum = _find_column(list(df.columns), "datum") or "Datum"
    col_zeit = _find_column(list(df.columns), "uhrzeit") or "Uhrzeit"
    col_ort = _find_column(list(df.columns), "entnahme") or "Entnahmeort"
    col_ecoli = _find_column(list(df.columns), "e", "coli") or "E.coli"
    col_bwq = _find_column(list(df.columns), "badewasser") or "Badewasserqualität"
    col_cm_ecoli = _find_column(list(df.columns), "coliminder", "e", "coli")
    col_cm_entero = _find_column(list(df.columns), "coliminder", "entero")

    dt = pd.to_datetime(
        df[col_datum].astype(str).str.strip() + " " + df[col_zeit].astype(str).str.strip(),
        format="%d.%m.%Y %H:%M:%S",
        errors="coerce",
    )
    mess = dt.dt.tz_localize(TZ, ambiguous="infer", nonexistent="shift_forward")

    out = pd.DataFrame(
        {
            "Messzeitpunkt_Labor": mess,
            "Entnahmeort": df[col_ort] if col_ort in df.columns else pd.NA,
            "E.coli": pd.to_numeric(df[col_ecoli], errors="coerce") if col_ecoli in df.columns else pd.NA,
            "Enterokokken": pd.NA,
            "Badewasserqualität": df[col_bwq] if col_bwq in df.columns else pd.NA,
            "coliminder_e_coli": pd.to_numeric(df[col_cm_ecoli], errors="coerce") if col_cm_ecoli else pd.NA,
            "coliminder_entero": pd.to_numeric(df[col_cm_entero], errors="coerce") if col_cm_entero else pd.NA,
        }
    )
    return out


def fetch_coliminder_long(
    client: OpenRemoteClient,
    *,
    from_date: date = EXPORT_FROM_DATE,
) -> pd.DataFrame:
    """Fetch Basel dashboard datapoints as a long DataFrame (same rows as CSV export)."""
    dashboards_payload = client.get_dashboards()
    basel = _find_dashboard_by_display_name(dashboards_payload, DASHBOARD_BASEL)
    if not basel:
        names = [
            str(x.get("displayName", ""))
            for x in _extract_items(dashboards_payload)
            if isinstance(x.get("displayName"), str)
        ]
        preview = ", ".join(sorted(set(names))[:25])
        raise RuntimeError(
            f"No dashboard with displayName {DASHBOARD_BASEL!r} found. "
            f"Known display names (sample): {preview or '(none)'}"
        )

    pairs = _basel_dashboard_attribute_pairs(basel)
    if not pairs:
        raise RuntimeError(
            f"Dashboard {DASHBOARD_BASEL!r} contains no attributeRefs for {', '.join(sorted(TARGET_ATTRIBUTE_NAMES))}."
        )

    LOGGER.info(
        "Basel dashboard: %s attribute source(s) for Ecoli/Entero/Transmission/Gehalt.",
        len(pairs),
    )

    assets_payload = client.query_assets()
    id_to_name = _build_asset_id_to_name(assets_payload)

    from_timestamp_ms = int(datetime.combine(from_date, datetime.min.time(), tzinfo=UTC).timestamp() * 1000)
    to_timestamp_ms = int(datetime.now(UTC).timestamp() * 1000)
    rows: list[dict[str, Any]] = []

    for pair in pairs:
        asset_id = pair["asset_id"]
        attribute_name = pair["attribute_name"]
        asset_name = id_to_name.get(asset_id, "")
        try:
            payload = client.get_datapoints(
                asset_id=asset_id,
                attribute_name=attribute_name,
                from_timestamp_ms=from_timestamp_ms,
                to_timestamp_ms=to_timestamp_ms,
            )
        except Exception as exc:  # pylint: disable=broad-exception-caught
            LOGGER.warning(
                "Failed to read datapoints for asset=%s attribute=%s: %s",
                asset_id,
                attribute_name,
                exc,
            )
            continue

        points = _normalize_datapoints(payload)
        for point in points:
            timestamp_ms = point.get("x", point.get("timestamp"))
            value = point.get("y", point.get("value"))
            timestamp_utc = ""
            if isinstance(timestamp_ms, (int, float)):
                timestamp_utc = datetime.fromtimestamp(timestamp_ms / 1000, tz=UTC).isoformat()
            rows.append(
                {
                    "asset_id": asset_id,
                    "asset_name": asset_name,
                    "attribute_name": attribute_name,
                    "timestamp_ms": timestamp_ms,
                    "timestamp_utc": timestamp_utc,
                    "value": value,
                }
            )

    rows.sort(key=lambda row: (row["asset_name"], row["attribute_name"], str(row["timestamp_ms"])))
    columns = ["asset_id", "asset_name", "attribute_name", "timestamp_ms", "timestamp_utc", "value"]
    return pd.DataFrame(rows, columns=columns)


def pivot_coliminder_wide(df_long: pd.DataFrame) -> pd.DataFrame:
    """Pivot long Coliminder rows to wide: Messzeitpunkt + Ecoli, Ecoli_2, Entero, Transmission, Gehalt."""
    if df_long.empty:
        return pd.DataFrame(columns=["Messzeitpunkt", *COLIMINDER_VALUE_COLUMNS])

    ts = pd.to_datetime(df_long["timestamp_ms"], unit="ms", utc=True, errors="coerce")
    mess = ts.dt.tz_convert(TZ)
    # Keep source time semantics but collapse millisecond jitter so Excel does not
    # show visually duplicated timestamps at second precision.
    df = df_long.assign(
        Messzeitpunkt=mess,
        MesszeitpunktSekunde=mess.dt.floor("s"),
    )
    df = df.dropna(subset=["Messzeitpunkt", "attribute_name"])

    wide = df.pivot_table(
        index="MesszeitpunktSekunde",
        columns="attribute_name",
        values="value",
        aggfunc="last",
    ).reindex(columns=COLIMINDER_VALUE_COLUMNS)
    wide = wide.reset_index().rename(columns={"MesszeitpunktSekunde": "Messzeitpunkt"}).sort_values("Messzeitpunkt")
    return wide


@dataclass(frozen=True)
class ValueChange:
    """A Coliminder value that the API reports differently than the persisted file."""

    timestamp: pd.Timestamp
    attribute: str
    old_value: Any
    new_value: Any


@dataclass(frozen=True)
class RemovedValue:
    """A Coliminder value the API no longer returns but that we keep on disk."""

    timestamp: pd.Timestamp
    attribute: str
    old_value: Any


@dataclass(frozen=True)
class MergeResult:
    wide: pd.DataFrame
    changes: list["ValueChange"]
    removed: list["RemovedValue"]


def _coerce_messzeitpunkt(series: pd.Series) -> pd.Series:
    """Return a tz-aware (Europe/Zurich), second-floored Messzeitpunkt series."""
    ts = pd.to_datetime(series, errors="coerce")
    if getattr(ts.dtype, "tz", None) is None:
        ts = ts.dt.tz_localize(TZ, ambiguous="infer", nonexistent="shift_forward")
    else:
        ts = ts.dt.tz_convert(TZ)
    return ts.dt.floor("s")


def load_existing_coliminder_wide(path: Path | str = COLIMINDER_XLSX) -> pd.DataFrame:
    """
    Load the previously persisted ``Coliminder.xlsx`` so we can merge instead of overwrite.

    The persisted file stores timezone-naive Europe/Zurich wall time; re-localize it so the
    timestamps line up with the freshly fetched, tz-aware frame. Missing files or unreadable
    content yield an empty frame (the merge then behaves like a first run).
    """
    path = Path(path)
    empty = pd.DataFrame(columns=["Messzeitpunkt", *COLIMINDER_VALUE_COLUMNS])
    if not path.exists():
        return empty
    try:
        df = pd.read_excel(path)
    except (OSError, ValueError, zipfile.BadZipFile, KeyError) as exc:
        LOGGER.warning("Existing %s could not be read (%s); treating as empty.", path, exc)
        return empty

    df.columns = [str(c).strip() for c in df.columns]
    if "Messzeitpunkt" not in df.columns:
        LOGGER.warning("Existing %s has no 'Messzeitpunkt' column; treating as empty.", path)
        return empty

    df["Messzeitpunkt"] = _coerce_messzeitpunkt(df["Messzeitpunkt"])
    df = df.dropna(subset=["Messzeitpunkt"])
    for col in COLIMINDER_VALUE_COLUMNS:
        df[col] = pd.to_numeric(df[col], errors="coerce") if col in df.columns else pd.NA
    return df[["Messzeitpunkt", *COLIMINDER_VALUE_COLUMNS]]


def _values_differ(old_value: Any, new_value: Any) -> bool:
    """True only when both values are present and meaningfully different.

    Adding a brand-new value (old missing) or losing one (new missing) is handled separately,
    so this returns ``False`` in those cases to avoid noisy "changed" notifications.
    """
    old_na = pd.isna(old_value)
    new_na = pd.isna(new_value)
    if old_na or new_na:
        return False
    try:
        return not math.isclose(float(old_value), float(new_value), rel_tol=1e-9, abs_tol=1e-9)
    except (TypeError, ValueError):
        return str(old_value) != str(new_value)


def merge_coliminder_wide(
    existing: pd.DataFrame,
    fetched: pd.DataFrame,
    *,
    from_date: date,
) -> MergeResult:
    """
    Merge freshly fetched Coliminder data into the persisted history without ever losing data.

    Rules:
    - Every timestamp/attribute value that was on disk is kept, even if the API no longer returns it.
    - When the API provides a value, it wins over the stored one (the source of truth for live data).
    - A value present in both but different is recorded as a change (-> e-mail).
    - A value that existed on disk inside the fetch window but is now missing from the API is recorded
      as a removed value that we deliberately preserve (-> e-mail), since this is exactly the data-loss
      situation we want to be warned about.
    """
    value_cols = COLIMINDER_VALUE_COLUMNS

    def _indexed(df: pd.DataFrame) -> pd.DataFrame:
        if df.empty:
            return pd.DataFrame(columns=value_cols, index=pd.DatetimeIndex([], tz=TZ, name="Messzeitpunkt"))
        tmp = df.copy()
        tmp["Messzeitpunkt"] = _coerce_messzeitpunkt(tmp["Messzeitpunkt"])
        tmp = tmp.dropna(subset=["Messzeitpunkt"]).set_index("Messzeitpunkt")
        tmp = tmp[~tmp.index.duplicated(keep="last")]
        for col in value_cols:
            if col not in tmp.columns:
                tmp[col] = pd.NA
        return tmp[value_cols]

    old = _indexed(existing)
    new = _indexed(fetched)
    all_index = old.index.union(new.index).sort_values()

    cutoff = pd.Timestamp(datetime.combine(from_date, datetime.min.time()), tz=TZ)
    in_window = all_index >= cutoff

    merged = pd.DataFrame(index=all_index)
    changes: list[ValueChange] = []
    removed: list[RemovedValue] = []

    for col in value_cols:
        old_col = old[col].reindex(all_index)
        new_col = new[col].reindex(all_index)
        # API value wins where present; otherwise keep what we already had.
        merged[col] = new_col.where(new_col.notna(), old_col)

        changed_mask = (old_col.notna() & new_col.notna()).to_numpy()
        for ts in all_index[changed_mask]:
            if _values_differ(old_col[ts], new_col[ts]):
                changes.append(ValueChange(ts, col, old_col[ts], new_col[ts]))

        removed_mask = (old_col.notna() & new_col.isna()).to_numpy() & in_window
        for ts in all_index[removed_mask]:
            removed.append(RemovedValue(ts, col, old_col[ts]))

    merged = merged.reset_index().rename(columns={"index": "Messzeitpunkt"})
    if "Messzeitpunkt" not in merged.columns and "level_0" in merged.columns:
        merged = merged.rename(columns={"level_0": "Messzeitpunkt"})
    merged = merged.sort_values("Messzeitpunkt").reset_index(drop=True)
    return MergeResult(wide=merged, changes=changes, removed=removed)


def _format_ts(ts: pd.Timestamp) -> str:
    try:
        return pd.Timestamp(ts).tz_convert(TZ).strftime("%d.%m.%Y %H:%M:%S")
    except (TypeError, ValueError):
        return str(ts)


def _format_value(value: Any) -> str:
    if pd.isna(value):
        return "(leer)"
    try:
        number = float(value)
    except (TypeError, ValueError):
        return str(value)
    if number == int(number):
        return str(int(number))
    return f"{number:g}"


def build_change_email_text(changes: list[ValueChange], removed: list[RemovedValue]) -> str:
    """Human-readable German notification body, mirroring the staka_kantonsblatt style."""
    text = (
        "Beim ETL-Lauf für die Coliminder-Daten (Badewasserqualität, Dataset 100530) "
        "wurden Abweichungen zwischen der API und den bereits gespeicherten Daten festgestellt.\n\n"
        "Die gespeicherten Daten werden NICHT gelöscht: Werte aus der API werden ergänzt bzw. "
        "aktualisiert, bereits vorhandene Werte bleiben erhalten.\n"
    )

    if changes:
        text += f"\nGeänderte Werte (API weicht vom gespeicherten Wert ab) – {len(changes)} Stück:\n"
        for change in changes[:200]:
            text += (
                f" - {_format_ts(change.timestamp)} | {change.attribute}: "
                f"{_format_value(change.old_value)} -> {_format_value(change.new_value)}\n"
            )
        if len(changes) > 200:
            text += f" - ... und {len(changes) - 200} weitere.\n"

    if removed:
        text += (
            f"\nWerte, welche die API nicht mehr liefert, aber bewusst erhalten bleiben "
            f"– {len(removed)} Stück:\n"
        )
        for item in removed[:200]:
            text += (
                f" - {_format_ts(item.timestamp)} | {item.attribute}: "
                f"{_format_value(item.old_value)} (in der API nicht mehr vorhanden)\n"
            )
        if len(removed) > 200:
            text += f" - ... und {len(removed) - 200} weitere.\n"

    text += f"\nDie Daten liegen hier:\n {COLIMINDER_XLSX}\n"
    text += "\nFreundliche Grüsse, \nEuer automatisierter Open Data Basel-Stadt Python Job"
    return text


def notify_coliminder_changes(changes: list[ValueChange], removed: list[RemovedValue]) -> None:
    """Send a notification e-mail when the API changed or dropped values. Never fails the ETL."""
    if not changes and not removed:
        LOGGER.info("No Coliminder value changes or removals detected; no e-mail sent.")
        return

    LOGGER.info(
        "Detected %s changed and %s removed Coliminder value(s); sending notification e-mail.",
        len(changes),
        len(removed),
    )
    text = build_change_email_text(changes, removed)
    try:
        msg = common.email_message(
            subject="Coliminder (100530): API hat Werte geändert/gelöscht – Daten wurden bewahrt.",
            text=text,
            img=None,
            attachment=None,
        )
        common.send_email(msg)
        LOGGER.info("Change-notification e-mail sent.")
    except Exception as exc:  # pylint: disable=broad-exception-caught
        LOGGER.warning("Could not send Coliminder change-notification e-mail: %s", exc)


def _is_zip_xlsx(content: bytes) -> bool:
    return len(content) >= 4 and content[:2] == b"PK"


def _try_read_xlsx_bytes(content: bytes) -> pd.DataFrame | None:
    try:
        return pd.read_excel(io.BytesIO(content))
    except (OSError, ValueError, zipfile.BadZipFile, KeyError):
        return None


def download_huwise_xlsx(
    ods_id: str,
    *,
    qv1: str | None,
    oldest_iso: str,
    time_field: str,
) -> pd.DataFrame:
    """
    Download Huwise / ODS explore XLSX export; save under ``data/{ods_id}.xlsx``.

    Tries ``qv1`` filter first; on failure or invalid payload, retries unfiltered
    and filters client-side on ``time_field``.
    """
    url = f"https://data.bs.ch/api/explore/v2.1/catalog/datasets/{ods_id}/exports/xlsx"
    params_base: dict[str, str] = {
        "lang": "de",
        "timezone": "Europe/Zurich",
        "use_labels": "true",
    }
    dest = OGD_DIR / f"{ods_id}.xlsx"

    qv1_effective = qv1 or f'{time_field} >= "{oldest_iso}"'
    resp = common.requests_get(url, params={**params_base, "qv1": qv1_effective})
    content = resp.content
    df_full: pd.DataFrame | None = None
    bytes_to_save: bytes | None = None

    if resp.ok and _is_zip_xlsx(content):
        df_full = _try_read_xlsx_bytes(content)
        if df_full is not None:
            bytes_to_save = content

    if df_full is None:
        LOGGER.warning(
            "Huwise %s: qv1 export not usable or not parseable (status=%s, len=%s). Retrying without qv1.",
            ods_id,
            getattr(resp, "status_code", "?"),
            len(content),
        )
        resp2 = common.requests_get(url, params=params_base)
        content2 = resp2.content
        if not resp2.ok or not _is_zip_xlsx(content2):
            raise RuntimeError(
                f"Huwise dataset {ods_id}: invalid XLSX response "
                f"(status={getattr(resp2, 'status_code', None)}, len={len(content2)})."
            )
        df_full = _try_read_xlsx_bytes(content2)
        if df_full is None:
            raise RuntimeError(f"Huwise dataset {ods_id}: could not parse XLSX after unfiltered download.")
        df_full = _filter_huwise_from_date(df_full, oldest_iso, time_field=time_field)
        buf = io.BytesIO()
        df_full.to_excel(buf, index=False)
        bytes_to_save = buf.getvalue()

    assert df_full is not None and bytes_to_save is not None
    dest.write_bytes(bytes_to_save)

    LOGGER.info("Huwise %s: wrote %s rows to %s", ods_id, len(df_full), dest)
    return df_full


def _filter_huwise_from_date(df: pd.DataFrame, oldest_iso: str, *, time_field: str) -> pd.DataFrame:
    """Client-side filter when API qv1 uses technical field names on label export."""
    cols_lower = {str(c).lower(): c for c in df.columns}
    candidates = [
        cols_lower.get(time_field.lower()),
        cols_lower.get("timestamp"),
        cols_lower.get("zeitstempel"),
        cols_lower.get("ende"),
        cols_lower.get("start"),
    ]
    time_col = next((c for c in candidates if c is not None), None)
    if time_col is None:
        LOGGER.warning("Could not detect time column for filter; returning full Huwise frame.")
        return df
    ts = pd.to_datetime(df[time_col], utc=True, errors="coerce").dt.tz_convert(TZ)
    cutoff = datetime.fromisoformat(oldest_iso).replace(tzinfo=TZ)
    if isinstance(cutoff, datetime) and cutoff.time() == datetime.min.time():
        pass
    mask = ts >= pd.Timestamp(cutoff)
    return df.loc[mask].reset_index(drop=True)


def normalize_huwise_100089_100271(df: pd.DataFrame) -> pd.DataFrame:
    """Timestamp-based Rhein frames: Messzeitpunkt + Wasserstand + Abflussmenge."""
    col_ts = _find_column(list(df.columns), "timestamp")
    if col_ts is None:
        col_ts = _find_column(list(df.columns), "zeitstempel") or "Zeitstempel"
    col_ws = _find_column(list(df.columns), "wasserstand")
    col_ab = _find_column(list(df.columns), "abfluss") or "Abflussmenge"
    if col_ws is None:
        col_ws = "Wasserstand"
    if col_ab is None:
        col_ab = "Abflussmenge"

    ts = pd.to_datetime(df[col_ts], utc=True, errors="coerce").dt.tz_convert(TZ)
    try:
        ts = ts.dt.as_unit("ns")
    except (AttributeError, TypeError):
        ts = ts.astype("datetime64[ns, Europe/Zurich]")
    out = pd.DataFrame(
        {
            "Messzeitpunkt": ts,
            "Wasserstand": pd.to_numeric(df[col_ws], errors="coerce") if col_ws in df.columns else pd.NA,
            "Abflussmenge": pd.to_numeric(df[col_ab], errors="coerce") if col_ab in df.columns else pd.NA,
        }
    )
    return out.dropna(subset=["Messzeitpunkt"]).sort_values("Messzeitpunkt").reset_index(drop=True)


def normalize_huwise_100323(df: pd.DataFrame) -> pd.DataFrame:
    """Trübung hourly: use interval end as Messzeitpunkt (matches prior R script)."""
    col_ende = _find_column(list(df.columns), "ende") or "Ende"
    col_tr = _find_column(list(df.columns), "trubung", "fnu") or _find_column(list(df.columns), "trubung")
    if col_tr is None:
        for c in df.columns:
            if "tr" in str(c).lower() and "fnu" in str(c).lower():
                col_tr = str(c)
                break
    if col_tr is None:
        raise ValueError("Could not find Trübung column in 100323 export.")

    ts = pd.to_datetime(df[col_ende], utc=True, errors="coerce").dt.tz_convert(TZ)
    try:
        ts = ts.dt.as_unit("ns")
    except (AttributeError, TypeError):
        ts = ts.astype("datetime64[ns, Europe/Zurich]")
    out = pd.DataFrame(
        {
            "Messzeitpunkt": ts,
            "Trübung": pd.to_numeric(df[col_tr], errors="coerce"),
        }
    )
    return out.dropna(subset=["Messzeitpunkt"]).sort_values("Messzeitpunkt").reset_index(drop=True)


def normalize_huwise_100046(df: pd.DataFrame) -> pd.DataFrame:
    """100046 water quality stream: temperature, oxygen, pH."""
    col_ts = _find_column(list(df.columns), "endezeitpunkt") or _find_column(list(df.columns), "startzeitpunkt")
    if col_ts is None:
        col_ts = "Endezeitpunkt"
    col_temp = _find_column(list(df.columns), "temperatur")
    col_o2 = _find_column(list(df.columns), "sauerstoffgehalt")
    col_ph = _find_column(list(df.columns), "ph")

    ts = pd.to_datetime(df[col_ts], utc=True, errors="coerce").dt.tz_convert(TZ)
    try:
        ts = ts.dt.as_unit("ns")
    except (AttributeError, TypeError):
        ts = ts.astype("datetime64[ns, Europe/Zurich]")

    out = pd.DataFrame(
        {
            "Messzeitpunkt": ts,
            "Wassertemperatur": pd.to_numeric(df[col_temp], errors="coerce") if col_temp else pd.NA,
            "Sauerstoffgehalt": pd.to_numeric(df[col_o2], errors="coerce") if col_o2 else pd.NA,
            "pH-Wert": pd.to_numeric(df[col_ph], errors="coerce") if col_ph else pd.NA,
        }
    )
    return out.dropna(subset=["Messzeitpunkt"]).sort_values("Messzeitpunkt").reset_index(drop=True)


def normalize_huwise_100243(df: pd.DataFrame) -> pd.DataFrame:
    """100243 Klingentalfähre water level: Zeitstempel + Wasserstand."""
    col_ts = _find_column(list(df.columns), "zeitstempel") or "Zeitstempel"
    col_ws = _find_column(list(df.columns), "wasserstand") or "Wasserstand"
    ts = pd.to_datetime(df[col_ts], utc=True, errors="coerce").dt.tz_convert(TZ)
    try:
        ts = ts.dt.as_unit("ns")
    except (AttributeError, TypeError):
        ts = ts.astype("datetime64[ns, Europe/Zurich]")
    out = pd.DataFrame(
        {
            "Messzeitpunkt": ts,
            "wasserstand_klingentalfaehre": pd.to_numeric(df[col_ws], errors="coerce"),
        }
    )
    return out.dropna(subset=["Messzeitpunkt"]).sort_values("Messzeitpunkt").reset_index(drop=True)


def _normalize_ts_series(ser: pd.Series) -> pd.Series:
    """``merge_asof`` requires identical datetime resolution / dtype for join keys."""
    out = pd.to_datetime(ser, utc=False, errors="coerce")
    if getattr(out.dtype, "tz", None) is not None:
        out = out.dt.tz_convert(TZ)
    else:
        out = out.dt.tz_localize(TZ, ambiguous="infer", nonexistent="shift_forward")
    try:
        return out.dt.as_unit("ns")
    except (AttributeError, TypeError):
        return out.astype("datetime64[ns]")


def _strip_tz_for_excel(df: pd.DataFrame) -> pd.DataFrame:
    """Excel writers require timezone-naive datetimes; keep Europe/Zurich wall time."""
    out = df.copy()
    for col in out.columns:
        ser = out[col]
        if pd.api.types.is_datetime64_any_dtype(ser) and getattr(ser.dtype, "tz", None) is not None:
            out[col] = ser.dt.tz_convert(TZ).dt.tz_localize(None)
    return out


def _write_human_readable_excel(df: pd.DataFrame, path: Path) -> None:
    """
    Write Excel with stable human-readable formatting:
    - first row bold
    - all columns width 20
    - date format ``dd.mmmm yyyy hh:mm:ss`` for time columns
    """
    clean = _strip_tz_for_excel(df)
    clean.to_excel(path, index=False)

    wb = load_workbook(path)
    ws = wb.active

    # Header row bold
    for cell in ws[1]:
        cell.font = Font(bold=True)

    # Uniform column width (requested openpyxl units)
    for col_cells in ws.iter_cols(min_row=1, max_row=1):
        ws.column_dimensions[col_cells[0].column_letter].width = 26.64

    date_cols = {"Messzeitpunkt", "Messzeitpunkt_Labor"}
    for col_idx, cell in enumerate(ws[1], start=1):
        if str(cell.value) in date_cols:
            for row_idx in range(2, ws.max_row + 1):
                ws.cell(row=row_idx, column=col_idx).number_format = "dd.mmmm yyyy hh:mm:ss"

    wb.save(path)


def _write_joined_workbook(labor_df: pd.DataFrame, coliminder_df: pd.DataFrame, path: Path) -> None:
    """Write joined_data.xlsx with two sheets: Labor and Coliminder."""
    labor_clean = _strip_tz_for_excel(labor_df)
    coli_clean = _strip_tz_for_excel(coliminder_df)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        labor_clean.to_excel(writer, sheet_name="Labor", index=False)
        coli_clean.to_excel(writer, sheet_name="Coliminder", index=False)

    wb = load_workbook(path)
    date_cols = {"datetime", "messzeitpunkt_labor", "messzeitpunkt"}
    for ws in wb.worksheets:
        for cell in ws[1]:
            cell.font = Font(bold=True)
        for col_cells in ws.iter_cols(min_row=1, max_row=1):
            ws.column_dimensions[col_cells[0].column_letter].width = 26.64
        for col_idx, cell in enumerate(ws[1], start=1):
            if str(cell.value) in date_cols:
                for row_idx in range(2, ws.max_row + 1):
                    ws.cell(row=row_idx, column=col_idx).number_format = "dd.mmmm yyyy hh:mm:ss"
    wb.save(path)


def _dedupe_messzeitpunkt(df: pd.DataFrame) -> pd.DataFrame:
    """``merge_asof`` expects a unique, sorted time key on the right side."""
    if df.empty or "Messzeitpunkt" not in df.columns:
        return df
    return df.drop_duplicates(subset=["Messzeitpunkt"], keep="last").reset_index(drop=True)


def _merge_asof_nearest(
    left: pd.DataFrame,
    right: pd.DataFrame,
    *,
    right_cols: list[str],
    tolerance: pd.Timedelta | None = None,
) -> pd.DataFrame:
    """Attach nearest-in-time rows from ``right`` to ``left`` on ``Messzeitpunkt``."""
    if right.empty or not right_cols:
        return left
    use_cols = ["Messzeitpunkt", *[c for c in right_cols if c in right.columns]]
    if len(use_cols) <= 1:
        return left
    left = left.sort_values("Messzeitpunkt", na_position="last").reset_index(drop=True)
    r = _dedupe_messzeitpunkt(right[use_cols].sort_values("Messzeitpunkt").reset_index(drop=True))
    return pd.merge_asof(
        left,
        r,
        on="Messzeitpunkt",
        direction="nearest",
        tolerance=tolerance,
        suffixes=("", "_r"),
    )


def _attach_coliminder_per_measurement(
    base: pd.DataFrame,
    df_coli_wide: pd.DataFrame,
    *,
    tolerance: pd.Timedelta,
) -> pd.DataFrame:
    """Join each Coliminder measurement independently to nearest timestamp."""
    if df_coli_wide.empty:
        for c in (
            "coliminder_e_coli",
            "coliminder_e_coli_2",
            "coliminder_entero",
            "coliminder_transimission",
            "coliminder_gehalt",
        ):
            if c not in base.columns:
                base[c] = pd.NA
        return base

    coli = df_coli_wide.copy()
    coli["Messzeitpunkt"] = _normalize_ts_series(coli["Messzeitpunkt"])
    mappings = [
        ("Ecoli", "coliminder_e_coli"),
        ("Ecoli_2", "coliminder_e_coli_2"),
        ("Entero", "coliminder_entero"),
        ("Transmission", "coliminder_transimission"),
        ("Gehalt", "coliminder_gehalt"),
    ]
    out = base
    for src, dst in mappings:
        if src not in coli.columns:
            out[dst] = pd.NA
            continue
        tmp_col = f"__{dst}_nearest"
        right = coli[["Messzeitpunkt", src]].dropna(subset=[src]).rename(columns={src: tmp_col})
        if right.empty:
            if dst not in out.columns:
                out[dst] = pd.NA
            continue
        out = _merge_asof_nearest(out, right, right_cols=[tmp_col], tolerance=tolerance)
        if dst in out.columns:
            out[dst] = out[dst].where(out[dst].notna(), out[tmp_col])
        else:
            out[dst] = out[tmp_col]
        out = out.drop(columns=[tmp_col])
    return out


def build_joined_sheets(
    df_coli_wide: pd.DataFrame,
    df_lab: pd.DataFrame,
    df_323: pd.DataFrame,
    df_089: pd.DataFrame,
    df_046: pd.DataFrame,
    df_243: pd.DataFrame,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Build two sheets: Labor-frequency and Coliminder-frequency."""
    one_hour = pd.Timedelta(hours=1)

    # Labor sheet (frequency of lab measurements)
    labor = df_lab.copy()
    labor["Messzeitpunkt"] = _normalize_ts_series(labor["Messzeitpunkt_Labor"])
    labor = labor[labor["Messzeitpunkt"].notna()].copy()
    labor = (
        labor.sort_values("Messzeitpunkt")
        .drop_duplicates(subset=["Messzeitpunkt"], keep="first")
        .reset_index(drop=True)
    )

    labor = _merge_asof_nearest(
        labor,
        df_046[["Messzeitpunkt", "Wassertemperatur", "Sauerstoffgehalt", "pH-Wert"]],
        right_cols=["Wassertemperatur", "Sauerstoffgehalt", "pH-Wert"],
        tolerance=one_hour,
    )
    labor = _merge_asof_nearest(
        labor, df_089[["Messzeitpunkt", "Abflussmenge"]], right_cols=["Abflussmenge"], tolerance=one_hour
    )
    labor = _merge_asof_nearest(
        labor, df_089[["Messzeitpunkt", "Wasserstand"]], right_cols=["Wasserstand"], tolerance=one_hour
    )
    labor = _merge_asof_nearest(labor, df_323[["Messzeitpunkt", "Trübung"]], right_cols=["Trübung"], tolerance=one_hour)
    labor = _merge_asof_nearest(
        labor,
        df_243[["Messzeitpunkt", "wasserstand_klingentalfaehre"]],
        right_cols=["wasserstand_klingentalfaehre"],
        tolerance=one_hour,
    )
    labor["Pegel"] = labor.get("Wasserstand")
    labor = _attach_coliminder_per_measurement(labor, df_coli_wide, tolerance=pd.Timedelta(hours=3))

    if "Entnahmeort" in labor.columns:
        labor["Entnahmeort"] = labor["Entnahmeort"].astype("string").str.replace(r"^Rheinwasser\s+", "", regex=True)
    labor = labor.rename(
        columns={
            "Messzeitpunkt_Labor": "messzeitpunkt_labor",
            "Entnahmeort": "entnahmeort",
            "E.coli": "e_coli",
            "Enterokokken": "enterokokken",
            "Badewasserqualität": "badewasserqualitaet",
            "Abflussmenge": "abflussmenge",
            "Trübung": "truebung",
            "Pegel": "pegel",
            "pH-Wert": "ph_wert",
            "Sauerstoffgehalt": "sauerstoffgehalt",
            "Wassertemperatur": "wassertemperatur",
            "Wasserstand": "wasserstand",
        }
    )
    labor_cols = [
        "messzeitpunkt_labor",
        "entnahmeort",
        "e_coli",
        "enterokokken",
        "badewasserqualitaet",
        "coliminder_e_coli",
        "coliminder_e_coli_2",
        "coliminder_entero",
        "coliminder_transimission",
        "coliminder_gehalt",
        "abflussmenge",
        "truebung",
        "pegel",
        "ph_wert",
        "sauerstoffgehalt",
        "wassertemperatur",
        "wasserstand",
        "wasserstand_klingentalfaehre",
    ]
    for c in labor_cols:
        if c not in labor.columns:
            labor[c] = pd.NA
    labor = (
        labor[labor_cols].sort_values("messzeitpunkt_labor", ascending=False, na_position="last").reset_index(drop=True)
    )

    # Coliminder sheet (frequency of coliminder, no lab measurements)
    coli = df_coli_wide.copy()
    coli["Messzeitpunkt"] = _normalize_ts_series(coli["Messzeitpunkt"])
    coli = (
        coli.sort_values("Messzeitpunkt").drop_duplicates(subset=["Messzeitpunkt"], keep="last").reset_index(drop=True)
    )
    coli = _merge_asof_nearest(
        coli,
        df_046[["Messzeitpunkt", "Wassertemperatur", "Sauerstoffgehalt", "pH-Wert"]],
        right_cols=["Wassertemperatur", "Sauerstoffgehalt", "pH-Wert"],
        tolerance=one_hour,
    )
    coli = _merge_asof_nearest(
        coli, df_089[["Messzeitpunkt", "Abflussmenge"]], right_cols=["Abflussmenge"], tolerance=one_hour
    )
    coli = _merge_asof_nearest(
        coli, df_089[["Messzeitpunkt", "Wasserstand"]], right_cols=["Wasserstand"], tolerance=one_hour
    )
    coli = _merge_asof_nearest(coli, df_323[["Messzeitpunkt", "Trübung"]], right_cols=["Trübung"], tolerance=one_hour)
    coli = _merge_asof_nearest(
        coli,
        df_243[["Messzeitpunkt", "wasserstand_klingentalfaehre"]],
        right_cols=["wasserstand_klingentalfaehre"],
        tolerance=one_hour,
    )
    coli["Pegel"] = coli.get("Wasserstand")
    coli = coli.rename(
        columns={
            "Messzeitpunkt": "messzeitpunkt",
            "Ecoli": "coliminder_e_coli",
            "Ecoli_2": "coliminder_e_coli_2",
            "Entero": "coliminder_entero",
            "Transmission": "coliminder_transimission",
            "Gehalt": "coliminder_gehalt",
            "Abflussmenge": "abflussmenge",
            "Trübung": "truebung",
            "Pegel": "pegel",
            "pH-Wert": "ph_wert",
            "Sauerstoffgehalt": "sauerstoffgehalt",
            "Wassertemperatur": "wassertemperatur",
            "Wasserstand": "wasserstand",
        }
    )
    coli_cols = [
        "messzeitpunkt",
        "coliminder_e_coli",
        "coliminder_e_coli_2",
        "coliminder_entero",
        "coliminder_transimission",
        "coliminder_gehalt",
        "abflussmenge",
        "truebung",
        "pegel",
        "ph_wert",
        "sauerstoffgehalt",
        "wassertemperatur",
        "wasserstand",
        "wasserstand_klingentalfaehre",
    ]
    for c in coli_cols:
        if c not in coli.columns:
            coli[c] = pd.NA
    coli = coli[coli_cols]

    # Extend joined "Coliminder" sheet with 2025 values from legacy short data.
    legacy_cols = {"Messzeitpunkt_Labor", "coliminder_e_coli", "coliminder_entero"}
    if legacy_cols.issubset(set(df_lab.columns)):
        legacy = df_lab[list(legacy_cols)].copy()
        legacy["messzeitpunkt"] = _normalize_ts_series(legacy["Messzeitpunkt_Labor"])
        legacy = legacy[legacy["messzeitpunkt"].notna()].copy()
        legacy = legacy[pd.to_datetime(legacy["messzeitpunkt"]).dt.year == 2025].copy()
        legacy["coliminder_e_coli_2"] = pd.NA
        legacy["coliminder_transimission"] = pd.NA
        legacy["coliminder_gehalt"] = pd.NA
        legacy["abflussmenge"] = pd.NA
        legacy["truebung"] = pd.NA
        legacy["pegel"] = pd.NA
        legacy["ph_wert"] = pd.NA
        legacy["sauerstoffgehalt"] = pd.NA
        legacy["wassertemperatur"] = pd.NA
        legacy["wasserstand"] = pd.NA
        legacy["wasserstand_klingentalfaehre"] = pd.NA
        legacy = legacy[coli_cols]

        coli = pd.concat([coli, legacy], ignore_index=True)
        coli = (
            coli.sort_values("messzeitpunkt", ascending=False, na_position="last")
            .drop_duplicates(subset=["messzeitpunkt"], keep="first")
            .reset_index(drop=True)
        )

    coli = coli.sort_values("messzeitpunkt", ascending=False, na_position="last").reset_index(drop=True)
    return labor, coli


def oldest_date_iso(
    df_lab: pd.DataFrame,
    df_coli_wide: pd.DataFrame,
) -> str:
    """Earliest calendar date (ISO) across lab and Coliminder for Huwise ``qv1``."""
    parts: list[pd.Timestamp] = []
    if not df_lab.empty and df_lab["Messzeitpunkt_Labor"].notna().any():
        parts.append(pd.Timestamp(df_lab["Messzeitpunkt_Labor"].min()))
    if not df_coli_wide.empty and df_coli_wide["Messzeitpunkt"].notna().any():
        parts.append(pd.Timestamp(df_coli_wide["Messzeitpunkt"].min()))
    if not parts:
        return EXPORT_FROM_DATE.isoformat()
    return min(parts).tz_convert(TZ).date().isoformat()


def labor_oldest_date(df_lab: pd.DataFrame) -> date:
    """Oldest lab date used to widen Coliminder extraction window."""
    if df_lab.empty or not df_lab["Messzeitpunkt_Labor"].notna().any():
        return EXPORT_FROM_DATE
    ts = pd.Timestamp(df_lab["Messzeitpunkt_Labor"].min())
    if ts.tzinfo is None:
        ts = ts.tz_localize(TZ)
    else:
        ts = ts.tz_convert(TZ)
    return min(EXPORT_FROM_DATE, ts.date())


def main() -> None:
    logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
    LOGGER.info("Starting Kantonslabor + Coliminder + Huwise ETL.")

    df_lab_new = load_labor_excel(LAB_EXCEL)
    df_lab_short = load_legacy_short_excel(LEGACY_SHORT_XLSX)
    # R-script baseline: uses `data_ecoli_entero_short`, not `Daten_LIMS_komplett`.
    df_lab = pd.concat([df_lab_short, df_lab_new], ignore_index=True, sort=False)
    LOGGER.info(
        "Loaded lab rows (R-like): short=%s, new=%s, total=%s.",
        len(df_lab_short),
        len(df_lab_new),
        len(df_lab),
    )

    config = Config.from_env()
    client = OpenRemoteClient(config=config)
    client.authenticate()
    LOGGER.info("Token endpoint used: %s", client.token_endpoint_used)
    LOGGER.info("Datapoint window: from %s 00:00 UTC through now.", EXPORT_FROM_DATE.isoformat())

    coliminder_from = COLIMINDER_EXPORT_FROM
    df_long = fetch_coliminder_long(client, from_date=coliminder_from)
    df_fetched_wide = pivot_coliminder_wide(df_long)

    # Never overwrite blindly: merge the freshly fetched data into the persisted history so that
    # data the API may have dropped (as happened before 22 May) is preserved on disk.
    df_existing_wide = load_existing_coliminder_wide(COLIMINDER_XLSX)
    merge_result = merge_coliminder_wide(df_existing_wide, df_fetched_wide, from_date=coliminder_from)
    df_coli_wide = merge_result.wide
    LOGGER.info(
        "Coliminder merge: existing=%s rows, fetched=%s rows, merged=%s rows, "
        "changed values=%s, preserved (API-removed) values=%s.",
        len(df_existing_wide),
        len(df_fetched_wide),
        len(df_coli_wide),
        len(merge_result.changes),
        len(merge_result.removed),
    )

    _write_human_readable_excel(df_coli_wide, COLIMINDER_XLSX)
    LOGGER.info(
        "Wrote Coliminder wide to %s (%s rows, from %s).",
        COLIMINDER_XLSX,
        len(df_coli_wide),
        coliminder_from.isoformat(),
    )

    notify_coliminder_changes(merge_result.changes, merge_result.removed)

    oldest = oldest_date_iso(df_lab, df_coli_wide)
    LOGGER.info("Huwise download lower bound (date): %s", oldest)

    df_089_raw = download_huwise_xlsx(
        "100089",
        qv1=f'timestamp >= "{oldest}"',
        oldest_iso=oldest,
        time_field="timestamp",
    )
    df_323_raw = download_huwise_xlsx(
        "100323",
        qv1=f'startzeitpunkt >= "{oldest}"',
        oldest_iso=oldest,
        time_field="startzeitpunkt",
    )
    df_046_raw = download_huwise_xlsx(
        "100046",
        qv1=f'startzeitpunkt >= "{oldest}"',
        oldest_iso=oldest,
        time_field="startzeitpunkt",
    )
    df_243_raw = download_huwise_xlsx(
        "100243",
        qv1=f'zeitstempel >= "{oldest}"',
        oldest_iso=oldest,
        time_field="zeitstempel",
    )

    df_089 = normalize_huwise_100089_100271(df_089_raw)
    df_323 = normalize_huwise_100323(df_323_raw)
    df_046 = normalize_huwise_100046(df_046_raw)
    df_243 = normalize_huwise_100243(df_243_raw)

    labor_sheet, coliminder_sheet = build_joined_sheets(df_coli_wide, df_lab, df_323, df_089, df_046, df_243)
    _write_joined_workbook(labor_sheet, coliminder_sheet, JOINED_XLSX)
    LOGGER.info(
        "Wrote joined dataset workbook to %s (Labor=%s rows, Coliminder=%s rows).",
        JOINED_XLSX,
        len(labor_sheet),
        len(coliminder_sheet),
    )

    LOGGER.info("ETL run finished.")


if __name__ == "__main__":
    main()

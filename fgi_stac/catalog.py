"""HUWISE bindings (Excel), flat publish catalog, and metadata snapshots."""

from __future__ import annotations

import logging
from pathlib import Path
from typing import Any

import yaml
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from dataspot_api import dataspot_metadata
from dataspot_auth import DataspotAuth
from metadata import flatten_to_snapshot
from paths import BINDINGS_FILE, ORIG_CATALOG_FILE, ORIG_METADATA_LAST_PUSH_FILE
from schema_merge import geometa_stac_url
from util import clean, normalize_uuid, row_key
from yaml_io import dump_yaml

_TEMPLATE_ORDER = ("custom", "dcat", "dcat_ap_ch", "default", "internal")
_SNAPSHOT_TEMPLATES = frozenset(_TEMPLATE_ORDER)

DATASETS_SHEET = "Datasets"
METADATA_LEGEND_SHEET = "MetadataLegend"
DATASET_COLUMNS = [
    "stac_collection_id",
    "geo_dataset",
    "huwise_id",
    "dataspot_asset_url",
    "stac_url",
    "stac_browser_url",
    "mapbs_url",
]
COLUMN_WIDTHS = {
    "stac_collection_id": 15.0,
    "geo_dataset": 60.0,
    "huwise_id": 15.0,
    "dataspot_asset_url": 75.0,
    "stac_url": 100.0,
    "stac_browser_url": 75.0,
    "mapbs_url": 50.0,
}
METADATA_LEGEND_ROWS = [
    ("HUWISE ids", "Edit the huwise_id column on the Datasets sheet."),
    ("Metadata policy", "Editorial metadata lives in HUWISE; publish uses conservative overwrite + data_orig/publish_metadata_last_push.yaml."),
    ("Pipeline jobs", "sync_catalog → prepare_assets → publish (or uv run etl.py for all three)."),
    ("Machine catalog", "data_orig/publish_catalog.yaml — flat, active HUWISE datasets only; do not edit manually."),
]

_METADATA_TEMPLATES = ("default", "dcat", "custom", "internal")


def is_snapshot_field_key(key: str) -> bool:
    text = clean(key)
    if "." not in text:
        return False
    template, _field = text.split(".", 1)
    return template in _SNAPSHOT_TEMPLATES


def _snapshot_sort_key(field_key: str) -> tuple[int, str]:
    template, field = field_key.split(".", 1)
    try:
        template_index = _TEMPLATE_ORDER.index(template)
    except ValueError:
        template_index = len(_TEMPLATE_ORDER)
    return template_index, field


def _huwise_sort_key(huwise_id: Any) -> tuple[int, str]:
    text = clean(huwise_id)
    if text.isdigit():
        return (0, f"{int(text):020d}")
    return (1, text.lower())


def filter_snapshot_entry(entry: dict[str, Any]) -> dict[str, Any]:
    out: dict[str, Any] = {}
    for key, value in entry.items():
        text_key = str(key)
        if not is_snapshot_field_key(text_key):
            continue
        if isinstance(value, str) and not clean(value):
            continue
        if isinstance(value, list) and not value:
            continue
        out[text_key] = value
    return out


def order_snapshot_entry(entry: dict[str, Any]) -> dict[str, Any]:
    filtered = filter_snapshot_entry(entry)
    return {key: filtered[key] for key in sorted(filtered, key=_snapshot_sort_key)}


def sort_snapshot_document(document: dict[str, Any]) -> dict[int | str, dict[str, Any]]:
    ordered: dict[int | str, dict[str, Any]] = {}
    for huwise_id in sorted(document, key=_huwise_sort_key):
        entry = document[huwise_id]
        if not isinstance(entry, dict):
            continue
        key: int | str = int(clean(huwise_id)) if clean(huwise_id).isdigit() else clean(huwise_id)
        ordered[key] = order_snapshot_entry(entry)
    return ordered


def _is_empty_snapshot_value(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, str):
        return not clean(value)
    if isinstance(value, list):
        return len(value) == 0
    return False


def merge_snapshot_entries(*entries: dict[str, Any]) -> dict[str, Any]:
    merged: dict[str, Any] = {}
    for entry in entries:
        if not isinstance(entry, dict):
            continue
        merged.update(entry)
    return order_snapshot_entry(merged)


def merge_snapshot_fill_gaps(base: dict[str, Any], *fallbacks: dict[str, Any]) -> dict[str, Any]:
    merged = dict(base)
    for entry in fallbacks:
        if not isinstance(entry, dict):
            continue
        for key, value in entry.items():
            if not is_snapshot_field_key(str(key)):
                continue
            if key not in merged or _is_empty_snapshot_value(merged.get(key)):
                if not _is_empty_snapshot_value(value):
                    merged[key] = value
    return order_snapshot_entry(merged)


def load_metadata_snapshot_document(path: Path) -> dict[str, dict[str, Any]]:
    if not path.is_file():
        return {}
    raw = yaml.safe_load(path.read_text(encoding="utf-8"))
    if not isinstance(raw, dict):
        return {}
    document: dict[str, Any] = {}
    for huwise_id, entry in raw.items():
        ods_id = clean(huwise_id)
        if ods_id and isinstance(entry, dict):
            document[ods_id] = entry
    return sort_snapshot_document(document)


def write_metadata_snapshot_file(path: Path, document: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(dump_yaml(sort_snapshot_document(document)), encoding="utf-8")


def _dataspot_uuid_from_asset_url(url: str) -> str:
    text = clean(url).rstrip("/")
    if "/assets/" not in text:
        return ""
    return text.rsplit("/assets/", 1)[-1].lower()


def load_active_dataset_rows(path: Path = BINDINGS_FILE) -> list[dict[str, str]]:
    if not path.is_file():
        return []
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet_name = DATASETS_SHEET if DATASETS_SHEET in workbook.sheetnames else workbook.sheetnames[0]
    sheet = workbook[sheet_name]
    header = [clean(cell.value).lower() for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    try:
        stac_idx = header.index("stac_collection_id")
        geo_idx = header.index("geo_dataset")
        hw_idx = header.index("huwise_id")
        asset_idx = header.index("dataspot_asset_url")
        mapbs_idx = header.index("mapbs_url")
    except ValueError as exc:
        workbook.close()
        raise ValueError(
            f"{path}: sheet {sheet_name!r} needs stac_collection_id, geo_dataset, "
            "huwise_id, dataspot_asset_url, mapbs_url"
        ) from exc
    rows: list[dict[str, str]] = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        huwise_id = clean(row[hw_idx] if hw_idx < len(row) else "")
        if not huwise_id:
            continue
        asset_url = clean(row[asset_idx] if asset_idx < len(row) else "")
        rows.append(
            {
                "huwise_id": huwise_id,
                "stac_collection_id": clean(row[stac_idx] if stac_idx < len(row) else ""),
                "geo_dataset": clean(row[geo_idx] if geo_idx < len(row) else ""),
                "dataspot_dataset_id": _dataspot_uuid_from_asset_url(asset_url),
                "dataspot_asset_url": asset_url,
                "mapbs_url": clean(row[mapbs_idx] if mapbs_idx < len(row) else ""),
            }
        )
    workbook.close()
    return rows


def load_bindings(path: Path = BINDINGS_FILE) -> dict[tuple[str, str], str]:
    if not path.is_file():
        return {}
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet_name = DATASETS_SHEET if DATASETS_SHEET in workbook.sheetnames else workbook.sheetnames[0]
    sheet = workbook[sheet_name]
    header = [clean(cell.value).lower() for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    try:
        stac_idx = header.index("stac_collection_id")
        geo_idx = header.index("geo_dataset")
        hw_idx = header.index("huwise_id")
    except ValueError as exc:
        workbook.close()
        raise ValueError(
            f"{path}: sheet {sheet_name!r} needs stac_collection_id, geo_dataset, huwise_id columns"
        ) from exc
    out: dict[tuple[str, str], str] = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        stac_id = clean(row[stac_idx] if stac_idx < len(row) else "")
        geo = clean(row[geo_idx] if geo_idx < len(row) else "")
        hw = clean(row[hw_idx] if hw_idx < len(row) else "")
        if stac_id and geo and hw:
            out[row_key(stac_id, geo)] = hw
    workbook.close()
    return out


def validate_bindings(bindings: dict[tuple[str, str], str]) -> None:
    seen_hw: set[str] = set()
    for hw in bindings.values():
        if not hw:
            continue
        if hw in seen_hw:
            raise ValueError(f"Duplicate huwise_id in bindings: {hw}")
        seen_hw.add(hw)


def load_bindings_with_fallback(
    catalog: dict[str, Any],
    *,
    bindings_path: Path = BINDINGS_FILE,
) -> dict[tuple[str, str], str]:
    bindings = load_bindings(bindings_path)
    if bindings:
        validate_bindings(bindings)
        return bindings
    if not bindings_path.is_file():
        logging.warning(
            "Missing %s; no HUWISE bindings loaded. Run ETL to regenerate the workbook.",
            bindings_path,
        )
    return {}


def _dataset_rows_from_catalog(catalog: dict[str, Any]) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    for collection in catalog.get("datasets", []):
        if not isinstance(collection, dict):
            continue
        stac_id = clean(collection.get("stac_collection_id"))
        stac_browser = clean(collection.get("stac_browser_url"))
        mapbs_url = clean(collection.get("mapbs_url"))
        for geo in collection.get("geo_datasets", []):
            if not isinstance(geo, dict):
                continue
            ds_id = normalize_uuid(geo.get("dataspot_dataset_id"))
            if not ds_id:
                continue
            rows.append(
                {
                    "stac_collection_id": stac_id,
                    "geo_dataset": clean(geo.get("geo_dataset")),
                    "huwise_id": clean(geo.get("huwise_id")),
                    "dataspot_asset_url": clean(geo.get("dataspot_asset_url"))
                    or f"https://bs.dataspot.io/web/prod/assets/{ds_id}",
                    "stac_url": geometa_stac_url(stac_id, ds_id),
                    "stac_browser_url": stac_browser,
                    "mapbs_url": mapbs_url,
                    "_dataspot_dataset_id": ds_id,
                }
            )
    rows.sort(key=lambda item: (item["stac_collection_id"].lower(), item["geo_dataset"].lower()))
    return rows


def _existing_huwise_by_key(path: Path) -> dict[tuple[str, str], str]:
    if not path.is_file():
        return {}
    try:
        return load_bindings(path)
    except ValueError:
        return {}


def _apply_column_widths(sheet: Worksheet, columns: list[str]) -> None:
    for idx, name in enumerate(columns, start=1):
        width = COLUMN_WIDTHS.get(name)
        if width:
            sheet.column_dimensions[get_column_letter(idx)].width = width


def write_bindings_workbook(
    catalog: dict[str, Any],
    path: Path = BINDINGS_FILE,
    *,
    bindings: dict[tuple[str, str], str] | None = None,
) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if bindings is None:
        bindings = load_bindings_with_fallback(catalog, bindings_path=path)
    preserved = _existing_huwise_by_key(path)
    preserved.update(bindings)
    rows = _dataset_rows_from_catalog(catalog)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = DATASETS_SHEET
    sheet.append(DATASET_COLUMNS)
    for cell in sheet[1]:
        cell.font = Font(bold=True)
    for row in rows:
        key = row_key(row["stac_collection_id"], row["geo_dataset"])
        sheet.append(
            [
                row["stac_collection_id"],
                row["geo_dataset"],
                preserved.get(key, row["huwise_id"]),
                row["dataspot_asset_url"],
                row["stac_url"],
                row["stac_browser_url"],
                row["mapbs_url"],
            ]
        )
    _apply_column_widths(sheet, DATASET_COLUMNS)

    legend = workbook.create_sheet(METADATA_LEGEND_SHEET)
    legend.append(["Topic", "Description"])
    for topic, description in METADATA_LEGEND_ROWS:
        legend.append([topic, description])

    workbook.save(path)
    logging.info("Wrote HUWISE bindings workbook: %s (%s rows on %s)", path, len(rows), DATASETS_SHEET)


def build_flat_publish_catalog(
    nested_catalog: dict[str, Any],
    bindings: dict[tuple[str, str], str],
) -> dict[str, dict[str, Any]]:
    flat: dict[str, dict[str, Any]] = {}
    seen_hw: set[str] = set()
    for collection in nested_catalog.get("datasets", []):
        if not isinstance(collection, dict):
            continue
        stac_id = clean(collection.get("stac_collection_id"))
        for geo in collection.get("geo_datasets", []):
            if not isinstance(geo, dict):
                continue
            geo_name = clean(geo.get("geo_dataset"))
            huwise_id = bindings.get(row_key(stac_id, geo_name), "") or clean(geo.get("huwise_id"))
            if not huwise_id:
                continue
            if huwise_id in seen_hw:
                raise ValueError(f"Duplicate huwise_id in active catalog: {huwise_id}")
            seen_hw.add(huwise_id)
            geo_row = dict(geo)
            geo_row["huwise_id"] = huwise_id
            flat[huwise_id] = flatten_to_snapshot(geo_row, collection)
    return flat


def enrich_snapshot_from_dataspot(
    entry: dict[str, Any],
    dataspot_dataset_id: str,
    *,
    auth: DataspotAuth | None = None,
) -> dict[str, Any]:
    ds_id = clean(dataspot_dataset_id).lower()
    if not ds_id:
        return entry
    client = auth if auth is not None else DataspotAuth()
    meta = dataspot_metadata(client, ds_id)
    return merge_snapshot_fill_gaps(
        entry,
        {
            "dcat.created": meta.get("created", ""),
            "dcat.issued": meta.get("issued", ""),
            "dcat.accrualperiodicity": meta.get("accrualperiodicity", ""),
        },
    )


def write_flat_publish_catalog(
    nested_catalog: dict[str, Any],
    bindings: dict[tuple[str, str], str],
    path: Path = ORIG_CATALOG_FILE,
) -> dict[str, dict[str, Any]]:
    flat = build_flat_publish_catalog(nested_catalog, bindings)
    dataspot_by_huwise: dict[str, str] = {}
    for collection in nested_catalog.get("datasets", []):
        if not isinstance(collection, dict):
            continue
        for geo in collection.get("geo_datasets", []):
            if not isinstance(geo, dict):
                continue
            hw = clean(geo.get("huwise_id"))
            ds = clean(geo.get("dataspot_dataset_id")).lower()
            if hw and ds:
                dataspot_by_huwise[hw] = ds

    auth = DataspotAuth()
    last_push = load_metadata_snapshot_document(ORIG_METADATA_LAST_PUSH_FILE) if ORIG_METADATA_LAST_PUSH_FILE.is_file() else {}
    enriched: dict[str, dict[str, Any]] = {}
    for huwise_id, entry in flat.items():
        merged = merge_snapshot_fill_gaps(entry, last_push.get(huwise_id, {}))
        ds_id = dataspot_by_huwise.get(huwise_id, "")
        if ds_id:
            merged = enrich_snapshot_from_dataspot(merged, ds_id, auth=auth)
        enriched[huwise_id] = merged
    flat = enriched
    write_metadata_snapshot_file(path, flat)
    logging.info("Wrote active publish catalog: %s (%s datasets)", path, len(flat))
    return flat


def active_huwise_ids_from_bindings(bindings: dict[tuple[str, str], str]) -> set[str]:
    return {hw for hw in bindings.values() if clean(hw)}


def prune_metadata_snapshot(
    path: Path,
    active_huwise_ids: set[str],
    *,
    label: str = "metadata snapshot",
) -> int:
    if not path.is_file() or not active_huwise_ids:
        return 0
    raw = yaml.safe_load(path.read_text(encoding="utf-8"))
    if not isinstance(raw, dict):
        return 0
    removed = 0
    pruned: dict[str, Any] = {}
    for key, value in raw.items():
        ods_id = clean(str(key))
        if ods_id in active_huwise_ids:
            pruned[key] = value
        else:
            removed += 1
    if removed:
        write_metadata_snapshot_file(path, pruned)
        logging.info("Pruned %s removed huwise_id(s) from %s (%s)", removed, label, path)
    return removed


def prune_all_publish_artifacts(active_huwise_ids: set[str]) -> None:
    prune_metadata_snapshot(ORIG_METADATA_LAST_PUSH_FILE, active_huwise_ids)


def is_nested_catalog(payload: dict[str, Any]) -> bool:
    return isinstance(payload.get("datasets"), list)


def flat_entry_to_geo_item(entry: dict[str, Any], *, huwise_id: str = "") -> dict[str, Any]:
    from metadata import dataspot_uuid_from_snapshot

    metadata: dict[str, dict[str, Any]] = {name: {} for name in _METADATA_TEMPLATES}
    for key, value in entry.items():
        text_key = clean(key)
        if "." not in text_key:
            continue
        template, field = text_key.split(".", 1)
        if template in metadata:
            metadata[template][field] = value
    return {
        "huwise_id": clean(huwise_id),
        "dataspot_dataset_id": dataspot_uuid_from_snapshot(entry),
        "geo_dataset": "",
        "metadata": metadata,
    }


def existing_geo_by_dataspot_uuid(path: Path = ORIG_CATALOG_FILE) -> dict[str, dict[str, Any]]:
    if not path.is_file():
        return {}
    raw = yaml.safe_load(path.read_text(encoding="utf-8"))
    if not isinstance(raw, dict):
        return {}
    by_uuid: dict[str, dict[str, Any]] = {}
    if is_nested_catalog(raw):
        for collection in raw.get("datasets", []):
            if not isinstance(collection, dict):
                continue
            for geo in collection.get("geo_datasets", []):
                if not isinstance(geo, dict):
                    continue
                ds = clean(geo.get("dataspot_dataset_id")).lower()
                if ds:
                    by_uuid[ds] = geo
        return by_uuid
    from metadata import dataspot_uuid_from_snapshot

    for huwise_id, entry in raw.items():
        if not isinstance(entry, dict):
            continue
        ds = dataspot_uuid_from_snapshot(entry)
        if ds:
            by_uuid[ds] = flat_entry_to_geo_item(entry, huwise_id=clean(str(huwise_id)))
    return by_uuid


def load_flat_publish_catalog(path: Path = ORIG_CATALOG_FILE) -> dict[str, dict[str, Any]]:
    if not path.is_file():
        return {}
    raw = yaml.safe_load(path.read_text(encoding="utf-8"))
    if not isinstance(raw, dict):
        return {}
    if is_nested_catalog(raw):
        raise ValueError(
            f"{path} uses the legacy nested STAC layout. Run `uv run sync_catalog.py` "
            "to regenerate the flat active publish catalog."
        )
    document: dict[str, dict[str, Any]] = {}
    for key, value in raw.items():
        if isinstance(value, dict):
            ods_id = clean(str(key))
            if ods_id:
                document[ods_id] = value
    return {str(k): v for k, v in sort_snapshot_document(document).items()}


# Re-exports used by publish_dataset
__all__ = [
    "filter_snapshot_entry",
    "load_active_dataset_rows",
    "load_bindings_with_fallback",
    "load_flat_publish_catalog",
    "merge_snapshot_entries",
    "merge_snapshot_fill_gaps",
    "order_snapshot_entry",
    "prune_all_publish_artifacts",
    "write_bindings_workbook",
    "write_flat_publish_catalog",
    "write_metadata_snapshot_file",
]

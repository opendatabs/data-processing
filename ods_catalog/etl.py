import logging
import os
from io import StringIO

import common
import openpyxl
import pandas as pd
from common import ODS_API_KEY


def main():
    # Get the new (published) datasets from ODS
    url_new_datasets = "https://data.bs.ch/explore/dataset/100055/download/"
    params = {
        "format": "csv",
        "use_labels_for_header": "true",
        "refine.visibility": "domain",
        "refine.publishing_published": "True",
    }
    headers = {"Authorization": f"apikey {ODS_API_KEY}"}
    r = common.requests_get(url_new_datasets, params=params, headers=headers)
    r.raise_for_status()
    df = common.pandas_read_csv(StringIO(r.text), sep=";", dtype=str)
    # Push the new datasets to ODS
    path_export = os.path.join("data", "100057_ods_catalog_published.csv")

    df.to_csv(path_export, index=False)
    common.update_ftp_and_odsp(path_export, "opendatabs", "100057")

    new_col = ["Title", "Issued", "url_dataset"]
    df_list = df[new_col]
    df_sorted = df_list.sort_values(by="Issued", ascending=False)
    # Make sure the 'Issued' column is a date
    df_sorted["Issued"] = pd.to_datetime(
        df_sorted["Issued"], errors="coerce", format="%Y-%m-%d"
    )
    # Remove missing values
    df_sorted = df_sorted.dropna(subset=["Issued"])
    # Create Excel Workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Datensätze"

    current_month = None
    column_width = 100  # Wider column for better readability
    ws.column_dimensions["A"].width = column_width
    bold_font = openpyxl.styles.Font(bold=True)

    for _, row in df_sorted.iterrows():
        month_year = row["Issued"].strftime("%m. %Y")
        issued_date = row["Issued"].strftime("%d.%m.%Y")
        entry_title = f"{row['Title']} / {issued_date}"
        entry_url = row["url_dataset"]

        # Add month as heading
        if month_year != current_month:
            ws.append([month_year])
            ws.cell(row=ws.max_row, column=1).font = bold_font  # Fett formatieren
            current_month = month_year

        # Add Title + Date
        ws.append([entry_title])

        # Add URL
        ws.append([entry_url])

    path_export_xlsx = os.path.join("data", "Datensatzliste.xlsx")
    wb.save(path_export_xlsx)


if __name__ == "__main__":
    logging.basicConfig(level=logging.DEBUG)
    logging.info(f"Executing {__file__}...")
    main()

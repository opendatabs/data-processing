import logging
import os

import common
import markdown
import pandas as pd
from markdown_newtab import NewTabExtension
from openpyxl import load_workbook


def cleanup_text_for_display(text):
    """
    Remove http://, https://, and www. from the text (if present).
    """
    if not text:
        return ""
    t = str(text)
    t = t.replace("https://", "").replace("http://", "").replace("www.", "")
    return t.strip()


def extract_link_and_text(cell):
    """
    Given an openpyxl cell, return (link, link_display_text) following these rules:
      1. If cell has a hyperlink, use that as 'link', and cell.value as 'link_display_text'
         (then cleanup link_display_text).
      2. If cell has no hyperlink but has text:
          - If that text starts with http:// or https://,
            use it for both link and link_display_text (then cleanup link_display_text).
          - Otherwise, prepend 'https://' to the text for 'link'
            and use the original text for link_display_text (then cleanup link_display_text).
      3. If empty, return ('', '').
    """
    if cell.value is None:
        # Nothing in the cell
        return "", ""

    if cell.hyperlink is not None:
        base_url = cell.hyperlink.target or ""
        fragment = cell.hyperlink.location or ""

        # Ensure https
        base_url = base_url.replace("http://", "https://")

        # Append location if present
        if fragment:
            if not fragment.startswith("#"):
                fragment = f"#{fragment}"
            real_url = f"{base_url}{fragment}"
        else:
            real_url = base_url

        displayed_text = cleanup_text_for_display(cell.value)
        return real_url, displayed_text

    else:
        # Cell has no hyperlink object, treat the value as plain text
        text = str(cell.value).strip()
        if text.lower().startswith("http://") or text.lower().startswith("https://"):
            link = text
            link = link.replace("http://", "https://")  # Ensure https
            displayed_text = cleanup_text_for_display(text)
            return link, displayed_text
        else:
            # Prepend https:// for the link, keep the original text as displayed text
            link = f"https://{text}"
            displayed_text = cleanup_text_for_display(text)
            return link, displayed_text


def main():
    # Iterate over every excel
    df_all = pd.DataFrame()
    for filename in os.listdir("data_orig"):
        if not filename.endswith(".xlsx"):
            logging.info(f"Ignoring {filename}; Not an Excel file.")
            continue
        if filename.startswith("~$"):
            logging.info(f"Ignoring {filename}; Temporary file.")
            continue
        logging.info(f"Processing {filename}...")
        excel_file_path = os.path.join("data_orig", filename)

        wb = load_workbook(excel_file_path, data_only=False)
        ws = wb.active  # or wb[sheetname] if you have a specific sheet

        df = pd.read_excel(excel_file_path, usecols="A:J", engine="openpyxl")
        df = df.rename(columns=lambda x: "Order" if x.startswith("Order") else x)
        df = df.rename(columns=lambda x: "Frage" if x.startswith("Frage") else x)
        df = df.rename(columns=lambda x: "Antwort" if x.startswith("Antwort") else x)
        df = df.rename(columns=lambda x: "Sprache" if x.startswith("Sprache") else x)
        df = df.rename(columns=lambda x: "Verantwortung" if x.startswith("Verantwortung") else x)
        df = df.rename(columns=lambda x: "Kontakt" if x.startswith("Kontakt") else x)
        df = df.rename(columns=lambda x: "Link Anzeigetext" if x.startswith("Link") else x)
        df = df.rename(columns=lambda x: "Zuletzt aktualisiert" if x.startswith("Zuletzt aktualisiert") else x)
        df = df.rename(columns=lambda x: "Thema" if x.startswith("Thema") else x)
        df = df.rename(columns=lambda x: "Keywords" if x.startswith("Keywords") else x)

        link_list = []
        link_text_list = []
        number_link_column = df.columns.get_loc("Link Anzeigetext")

        for row_idx, row in enumerate(
            ws.iter_rows(min_row=2, max_row=1 + df.shape[0], min_col=1, max_col=10),
            start=2,
        ):
            cell_link = row[number_link_column]
            link_val, text_val = extract_link_and_text(cell_link)
            link_list.append(link_val)
            link_text_list.append(text_val)

        df["Link"] = link_list
        df["Link Anzeigetext"] = link_text_list

        logging.info(f"Processing {filename} with {df.shape[0]} rows. Turning markdown into HTML...")

        df["Antwort HTML"] = df["Antwort"].apply(
            lambda x: markdown.markdown(x, extensions=["nl2br", NewTabExtension()]) if pd.notna(x) else x
        )

        df_all = pd.concat([df_all, df], ignore_index=True)

    if df_all.empty:
        logging.error("No data found. Exiting...")
        return

    path_export = os.path.join("data", "100443_dcc_ki_faq.csv")
    df_all.to_csv(path_export, index=False)
    common.update_ftp_and_odsp(
        path_export=path_export,
        folder_name="dcc/ki",
        dataset_id="100443",
    )


if __name__ == "__main__":
    logging.basicConfig(level=logging.DEBUG)
    logging.info(f"Executing {__file__}...")
    main()
    logging.info("Job completed successfully!")

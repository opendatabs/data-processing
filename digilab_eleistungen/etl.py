import logging
import os
from collections import defaultdict
from io import StringIO

import common
import pandas as pd
from dotenv import load_dotenv
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from rapidfuzz import fuzz, process as fuzz_process, utils as fuzz_utils

load_dotenv()

ODS_API_KEY = os.getenv("ODS_API_KEY")

FILL_GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
FILL_YELLOW = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
FILL_RED = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
FILL_HEADER = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
FONT_HEADER = Font(bold=True, color="FFFFFF")


DEPT_ABBREVIATIONS: dict[str, str] = {
    "PD": "Präsidialdepartement",
    "BVD": "Bau- und Verkehrsdepartement",
    "ED": "Erziehungsdepartement",
    "FD": "Finanzdepartement",
    "GD": "Gesundheitsdepartement",
    "JSD": "Justiz- und Sicherheitsdepartement",
    "WSU": "Departement für Wirtschaft, Soziales und Umwelt",
}

# Dienststellen per department as shown on page 2 of the official organigramm
# (bs-organigramme-2026.pdf, Stand 1. Januar 2026).
# Only these Ebene-3 organisations (plus any "Departementsleitung …" rows)
# are kept in the output; everything else (Kommissionen, historical entries,
# Baurekurskommission, etc.) is filtered out.
ORGANIGRAMM_DIENSTSTELLEN: dict[str, set[str]] = {
    "Präsidialdepartement": {
        "Generalsekretariat PD",
        "Fachstelle Klima",
        "Staatskanzlei",
        "Aussenbeziehungen und Standortmarketing",
        "Gleichstellung und Diversität",
        "Kantons- und Stadtentwicklung",
        "Abteilung Kultur",
        "Statistisches Amt",
    },
    "Bau- und Verkehrsdepartement": {
        "Generalsekretariat BVD",
        "Bau- und Gastgewerbeinspektorat",
        "Städtebau & Architektur",
        "Tiefbauamt",
        "Mobilität",
        "Grundbuch- und Vermessungsamt",
        "Stadtgärtnerei",
    },
    "Erziehungsdepartement": {
        "Generalsekretariat",
        "Volksschulen",
        "Mittelschulen und Berufsbildung",
        "Hochschulen",
        "Jugend, Familie und Sport",
        "Zentrale Dienste",
    },
    "Finanzdepartement": {
        "Generalsekretariat",
        "Finanzverwaltung",
        "Steuerverwaltung",
        "IT BS",
        "Human Resources Basel-Stadt",
        "Immobilien Basel-Stadt",
    },
    "Gesundheitsdepartement": {
        "Generalsekretariat",
        "Gesundheitsbeteiligungen und Finanzen",
        "Kommunikation",
        "Bereich Gesundheitsversorgung",
        "Institut für Rechtsmedizin",
        "Kantonales Laboratorium",
        "Kantonales Veterinäramt",
        "Medizinische Dienste",
        "Abteilung Sucht",
    },
    "Justiz- und Sicherheitsdepartement": {
        "Generalsekretariat",
        "Zentraler Rechtsdienst",
        "Kantonspolizei",
        "Rettung",
        "Bevölkerungsdienste und Migration",
    },
    "Departement für Wirtschaft, Soziales und Umwelt": {
        "Generalsekretariat",
        "Amt für Sozialbeiträge (ASB)",
        "Kindes- und Erwachsenenschutzbehörde",
        "Amt für Beistandschaften und Erwachsenenschutz ABES",
        "Sozialhilfe",
        "Amt für Wirtschaft und Arbeit (AWA)",
        "Amt für Umwelt und Energie (AUE)",
        "Amt für Wald und Wild beider Basel",
    },
}


def get_dataset(ods_id: str) -> pd.DataFrame:
    url = f"https://data.bs.ch/explore/dataset/{ods_id}/download/"
    params = {"format": "csv", "use_labels_for_header": "true"}
    headers = {"Authorization": f"apikey {ODS_API_KEY}"}
    r = common.requests_get(url, params=params, headers=headers)
    r.raise_for_status()
    return common.pandas_read_csv(StringIO(r.text), sep=";", dtype=str)


def compute_subtree_leistungen(df: pd.DataFrame) -> pd.Series:
    """For each org, sum up all Leistungen in its entire sub-tree via path prefixes."""
    total_counts: dict[str, int] = defaultdict(int)
    for _, row in df[df["Anzahl_Leistungen_direkt"] > 0].iterrows():
        parts = row["Organisationspfad"].split(">")
        count = row["Anzahl_Leistungen_direkt"]
        for i in range(len(parts)):
            total_counts[">".join(parts[: i + 1])] += count
    return df["Organisationspfad"].map(total_counts).fillna(0).astype(int)


def auto_fit_columns(ws, max_width=60):
    for col_idx, col_cells in enumerate(ws.columns, 1):
        max_len = 0
        for cell in col_cells:
            if cell.value:
                lines = str(cell.value).split("\n")
                max_len = max(max_len, max(len(line) for line in lines))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, max_width)


def format_header(ws):
    for cell in ws[1]:
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"


def color_code_rows(ws, direct_col_idx: int, total_col_idx: int):
    """Green = has direct Leistungen, Yellow = only sub-orgs have, Red = none at all."""
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        direct = row[direct_col_idx - 1].value or 0
        total = row[total_col_idx - 1].value or 0
        if total == 0:
            fill = FILL_RED
        elif direct > 0:
            fill = FILL_GREEN
        else:
            fill = FILL_YELLOW
        for cell in row:
            cell.fill = fill


def main():
    df_sk = get_dataset("100349")
    df_leistungen = get_dataset("100324")

    # --- Build hierarchy from Organisationspfad ---
    path_levels = df_sk["Organisationspfad"].str.split(">", expand=True)
    max_depth = path_levels.shape[1]
    level_cols = [f"Ebene {i + 1}" for i in range(max_depth)]
    path_levels.columns = level_cols
    df_sk = pd.concat([df_sk, path_levels], axis=1)
    df_sk["Tiefe"] = df_sk["Organisationspfad"].str.count(">") + 1

    # --- Fuzzy-match Dienststelle names to Staatskalender organisations ---
    all_sk_names = df_sk["Organisationsname"].tolist()

    dienststelle_to_sk_id: dict[str, str] = {}
    for dept_abbrev, dienststelle in (
        df_leistungen[["Departement", "Dienststelle"]]
        .drop_duplicates()
        .itertuples(index=False, name=None)
    ):
        dept_full = DEPT_ABBREVIATIONS.get(dept_abbrev)

        # Search within department subtree first, then also try the full list.
        # Prefer the dept-scoped result to avoid cross-department mismatches
        # (e.g. multiple "Generalsekretariat" entries), but fall back to
        # the global result when it's a perfect match -- this handles orgs
        # like "Staatsanwaltschaft" that sit at Ebene 2 outside any dept.
        dept_name, dept_score, dept_id = None, 0.0, None
        glob_name, glob_score, glob_id = None, 0.0, None

        if dept_full:
            mask = df_sk["Organisationspfad"].str.startswith(
                f"Regierung und Verwaltung>{dept_full}"
            )
            dept_candidates = df_sk[mask]
            dept_match = fuzz_process.extractOne(
                dienststelle,
                dept_candidates["Organisationsname"].tolist(),
                scorer=fuzz.WRatio,
                processor=fuzz_utils.default_process,
                score_cutoff=70.0,
            )
            if dept_match:
                dept_name, dept_score = dept_match[0], dept_match[1]
                dept_id = dept_candidates.iloc[dept_match[2]]["ID Organisation"]

        global_match = fuzz_process.extractOne(
            dienststelle,
            all_sk_names,
            scorer=fuzz.WRatio,
            processor=fuzz_utils.default_process,
            score_cutoff=70.0,
        )
        if global_match:
            glob_name, glob_score = global_match[0], global_match[1]
            glob_id = df_sk.iloc[global_match[2]]["ID Organisation"]

        if dept_name and (dept_score >= glob_score or glob_score < 100):
            best_name, best_score, best_id = dept_name, dept_score, dept_id
        else:
            best_name, best_score, best_id = glob_name, glob_score, glob_id

        if best_name:
            dienststelle_to_sk_id[dienststelle] = best_id
            if best_score < 100:
                logging.info(
                    f"Fuzzy match: '{dienststelle}' -> '{best_name}' "
                    f"(score: {best_score:.1f}, ID: {best_id})"
                )
        else:
            logging.warning(f"No match found for Dienststelle: '{dienststelle}'")

    df_leistungen["Matched_SK_ID"] = df_leistungen["Dienststelle"].map(
        dienststelle_to_sk_id
    )

    # --- Refine match using "Weitere Gliederung OE" sub-unit names ---
    # Try to match each sub-unit to a more specific child org within the
    # already-matched Dienststelle's subtree in the Staatskalender.
    skip_values = {"keine", "Keine", "Alle Abteilungen", "-"}
    sk_path_by_id = dict(
        zip(df_sk["ID Organisation"], df_sk["Organisationspfad"])
    )
    for dienststelle, sub_unit in (
        df_leistungen[["Dienststelle", "Weitere Gliederung OE"]]
        .drop_duplicates()
        .itertuples(index=False, name=None)
    ):
        if sub_unit in skip_values or pd.isna(sub_unit):
            continue
        parent_id = dienststelle_to_sk_id.get(dienststelle)
        if not parent_id:
            continue
        parent_path = sk_path_by_id.get(parent_id, "")
        subtree = df_sk[
            df_sk["Organisationspfad"].str.startswith(parent_path + ">")
        ]
        if subtree.empty:
            continue
        match = fuzz_process.extractOne(
            sub_unit,
            subtree["Organisationsname"].tolist(),
            scorer=fuzz.WRatio,
            processor=fuzz_utils.default_process,
            score_cutoff=80.0,
        )
        if match:
            matched_name, score, idx = match
            child_id = subtree.iloc[idx]["ID Organisation"]
            mask = (
                (df_leistungen["Dienststelle"] == dienststelle)
                & (df_leistungen["Weitere Gliederung OE"] == sub_unit)
            )
            df_leistungen.loc[mask, "Matched_SK_ID"] = child_id
            if score < 100:
                logging.info(
                    f"Sub-org match: '{sub_unit}' -> '{matched_name}' "
                    f"(score: {score:.1f}, ID: {child_id})"
                )

    unmatched_mask = df_leistungen["Matched_SK_ID"].isna()
    if unmatched_mask.any():
        unmatched_rows = (
            df_leistungen.loc[unmatched_mask, ["Dienststelle", "Weitere Gliederung OE"]]
            .drop_duplicates()
        )
        logging.warning(
            f"Dienststellen ohne Match im Staatskalender:\n{unmatched_rows.to_string(index=False)}"
        )

    MAX_EBENE = 4

    # --- Aggregate Leistungen per matched organisation ---
    leistungen_agg = (
        df_leistungen[~unmatched_mask]
        .groupby("Matched_SK_ID")
        .agg(
            Anzahl_Leistungen_direkt=("Leistungs-ID", "count"),
            Leistungsnamen=("Name", lambda x: "\n".join(sorted(x))),
        )
        .reset_index()
    )

    # --- Merge onto Staatskalender ---
    df = df_sk.merge(
        leistungen_agg,
        left_on="ID Organisation",
        right_on="Matched_SK_ID",
        how="left",
    )
    df["Anzahl_Leistungen_direkt"] = (
        df["Anzahl_Leistungen_direkt"].fillna(0).astype(int)
    )
    df["Anzahl_Leistungen_inkl_Unterorg"] = compute_subtree_leistungen(df)
    df["Hat_Leistungen"] = (
        (df["Anzahl_Leistungen_direkt"] > 0).map({True: "Ja", False: "Nein"})
    )

    # Keep only Dienststellen (Ebene 3) that appear in the official organigramm,
    # plus any "Departementsleitung …" rows.  Rows at Ebene 1/2 (no Ebene 3)
    # and rows under departments not listed in the whitelist pass through.
    allowed_pairs = {
        (dept, name)
        for dept, names in ORGANIGRAMM_DIENSTSTELLEN.items()
        for name in names
    }
    needs_check = df["Ebene 3"].notna() & df["Ebene 2"].isin(ORGANIGRAMM_DIENSTSTELLEN)
    is_allowed = pd.Series(
        [(e2, e3) in allowed_pairs for e2, e3 in zip(df["Ebene 2"], df["Ebene 3"])],
        index=df.index,
    )
    is_dept_lead = df["Ebene 3"].str.startswith("Departementsleitung", na=False)
    organigramm_mask = ~(needs_check & ~is_allowed & ~is_dept_lead)
    filtered_names = sorted(
        df.loc[~organigramm_mask, "Organisationsname"].unique()
    )
    if filtered_names:
        logging.info(
            f"Organigramm filter removed {len(filtered_names)} org names: "
            f"{filtered_names}"
        )
    df = df[organigramm_mask]

    # Drop entire Ebene-1 branches that carry no Leistungen at all
    ebene1_with_leistungen = set(
        df.loc[df["Anzahl_Leistungen_inkl_Unterorg"] > 0, "Ebene 1"].unique()
    )
    dropped = set(df["Ebene 1"].unique()) - ebene1_with_leistungen
    if dropped:
        logging.info(f"Dropping Ebene-1 branches without Leistungen: {dropped}")
    df = df[df["Ebene 1"].isin(ebene1_with_leistungen)]

    # Build "Weitere Ebenen" for orgs deeper than Ebene 4
    df["Weitere Ebenen"] = df["Organisationspfad"].apply(
        lambda p: " > ".join(p.split(">")[MAX_EBENE:]) if p.count(">") >= MAX_EBENE else ""
    )

    # Collapse branches whose children carry no Leistungen.  An org is only
    # "expandable" (its children shown) when at least one child has
    # Leistungen.  E.g. if Kantons- und Stadtentwicklung (Ebene 3) has
    # direct matches but none of its Ebene 4 sub-orgs do, only the Ebene 3
    # row is shown.  Same for Staatsanwaltschaft at Ebene 2.
    expandable_paths: set[str] = set()
    for path in df.loc[
        df["Anzahl_Leistungen_inkl_Unterorg"] > 0, "Organisationspfad"
    ]:
        parts = path.split(">")
        for i in range(1, len(parts)):
            expandable_paths.add(">".join(parts[:i]))
    keep_mask = pd.Series(True, index=df.index)
    for idx, row in df.iterrows():
        parts = row["Organisationspfad"].split(">")
        for i in range(1, len(parts)):
            ancestor_path = ">".join(parts[:i])
            if ancestor_path not in expandable_paths:
                keep_mask[idx] = False
                break
    df = df[keep_mask]
    df = df.sort_values("Organisationspfad").reset_index(drop=True)

    # ====== Sheet 1: Alle Organisationen (full hierarchy) ======
    display_level_cols = [f"Ebene {i + 1}" for i in range(MAX_EBENE)]
    all_org_cols = display_level_cols + [
        "Weitere Ebenen",
        "ID Organisation",
        "Organisationsname",
        "Tiefe",
        "Anzahl_Leistungen_direkt",
        "Anzahl_Leistungen_inkl_Unterorg",
        "Hat_Leistungen",
        "Leistungsnamen",
    ]
    df_all = df[all_org_cols].copy()

    # ====== Sheet 2: Zusammenfassung Departemente ======
    regierung = df[df["Ebene 1"] == "Regierung und Verwaltung"]

    # Include Staatsanwaltschaft (lives under "Weitere öffentliche Stellen")
    staatsanwaltschaft = df[df["Ebene 2"] == "Staatsanwaltschaft"]
    dept_scope = pd.concat([regierung, staatsanwaltschaft]).drop_duplicates()

    dept_rows = []
    for dept_name in sorted(dept_scope["Ebene 2"].dropna().unique()):
        dept_data = dept_scope[dept_scope["Ebene 2"] == dept_name]
        ebene3_rows = dept_data[dept_data["Tiefe"] == 3]
        gs_mask = ebene3_rows["Ebene 3"].str.startswith("Generalsekretariat")
        dl_mask = ebene3_rows["Ebene 3"].str.startswith("Departementsleitung")
        dienststellen = ebene3_rows[~gs_mask & ~dl_mask]
        dept_rows.append({
            "Departement": dept_name,
            "Anzahl_Dienststellen": len(dienststellen),
            "Dienststellen_mit_Leistungen": int(
                (dienststellen["Anzahl_Leistungen_inkl_Unterorg"] > 0).sum()
            ),
            "Dienststellen_ohne_Leistungen": int(
                (dienststellen["Anzahl_Leistungen_inkl_Unterorg"] == 0).sum()
            ),
            "Anzahl_Leistungen_Generalsekretariat": (
                int(ebene3_rows.loc[gs_mask, "Anzahl_Leistungen_inkl_Unterorg"].sum())
                if gs_mask.any() else 0
            ),
            "Anzahl_Leistungen_Departementsleitung": (
                int(ebene3_rows.loc[dl_mask, "Anzahl_Leistungen_inkl_Unterorg"].sum())
                if dl_mask.any() else 0
            ),
            "Total_Leistungen": int(dept_data["Anzahl_Leistungen_direkt"].sum()),
        })
    dept_summary = pd.DataFrame(dept_rows).sort_values(
        "Total_Leistungen", ascending=False
    )

    # ====== Sheet 3: Zusammenfassung Dienststellen ======
    has_ebene3 = dept_scope[dept_scope["Ebene 3"].notna()]
    amt_summary = (
        has_ebene3.groupby(["Ebene 2", "Ebene 3"])
        .agg(Total_Leistungen=("Anzahl_Leistungen_direkt", "sum"))
        .reset_index()
        .rename(columns={"Ebene 2": "Departement", "Ebene 3": "Dienststelle"})
        .sort_values(["Departement", "Total_Leistungen"], ascending=[True, False])
    )
    # Add Ebene-2 entities that have direct Leistungen but no Ebene-3 children
    # (e.g. Staatsanwaltschaft) as their own Dienststelle row.
    ebene2_direct = dept_scope[
        (dept_scope["Tiefe"] == 2) & (dept_scope["Anzahl_Leistungen_direkt"] > 0)
    ]
    extra_rows = [
        {
            "Departement": row["Ebene 2"],
            "Dienststelle": row["Ebene 2"],
            "Total_Leistungen": int(row["Anzahl_Leistungen_direkt"]),
        }
        for _, row in ebene2_direct.iterrows()
        if row["Ebene 2"] not in amt_summary["Departement"].values
    ]
    if extra_rows:
        amt_summary = pd.concat(
            [amt_summary, pd.DataFrame(extra_rows)], ignore_index=True
        )

    # ====== Write Excel ======
    output_path = os.path.join("data", "leistungen_uebersicht.xlsx")
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df_all.to_excel(writer, sheet_name="Alle Organisationen", index=False)
        dept_summary.to_excel(writer, sheet_name="Departemente", index=False)
        amt_summary.to_excel(writer, sheet_name="Dienststellen", index=False)
        df_leistungen.to_excel(writer, sheet_name="Leistungen Detail", index=False)

        ws_all = writer.sheets["Alle Organisationen"]
        direct_col = all_org_cols.index("Anzahl_Leistungen_direkt") + 1
        total_col = all_org_cols.index("Anzahl_Leistungen_inkl_Unterorg") + 1
        format_header(ws_all)
        color_code_rows(ws_all, direct_col, total_col)
        for row in ws_all.iter_rows(min_row=2, max_row=ws_all.max_row):
            leistung_cell = row[all_org_cols.index("Leistungsnamen")]
            leistung_cell.alignment = Alignment(wrap_text=True, vertical="top")
        auto_fit_columns(ws_all)

        for name in ["Departemente", "Dienststellen", "Leistungen Detail"]:
            ws = writer.sheets[name]
            format_header(ws)
            auto_fit_columns(ws)

    logging.info(f"Excel written to {output_path}")


if __name__ == "__main__":
    logging.basicConfig(level=logging.DEBUG)
    logging.info(f"Executing {__file__}...")
    main()
    logging.info("Job successful.")

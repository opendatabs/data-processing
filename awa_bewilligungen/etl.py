"""ETL pipeline for AWA Bewilligungen (work permits and shop opening hours).

Reads monthly indicator tables from an Excel workbook published by the
Arbeitsinspektorat and exports them as tidy CSV files.
"""

import logging
from datetime import date
from pathlib import Path

import common
import openpyxl
import pandas as pd
from dotenv import load_dotenv
from openpyxl.utils.cell import column_index_from_string, coordinate_from_string

load_dotenv()
logger = logging.getLogger(__name__)

SOURCE_FILE = Path("data_orig/Indikatoren - inkl. AMI 2015 - 2025.xlsx")
SHEET_NAME = "Arbeitsinspektorat"

MONTH_NAME_TO_NUMBER: dict[str, int] = {
    "Januar": 1,
    "Februar": 2,
    "März": 3,
    "April": 4,
    "Mai": 5,
    "Juni": 6,
    "Juli": 7,
    "August": 8,
    "September": 9,
    "Oktober": 10,
    "November": 11,
    "Dezember": 12,
}

DATASETS: list[dict[str, str]] = [
    {
        "name": "Ladenöffnungszeiten",
        "top_left_cell": "B118",
        "output": "data/100424_ladenoeffnungszeiten.csv",
        "dataset_id": "100424",
    },
    {
        "name": "Arbeitszeitbewilligungen",
        "top_left_cell": "B29",
        "output": "data/100425_arbeitszeitbewilligungen.csv",
        "dataset_id": "100425",
    },
]


def read_monthly_table(filename: Path, sheet_name: str, top_left_cell: str) -> pd.DataFrame:
    """Read a year-by-month table from an Excel sheet and return it in long format.

    The table layout has year headers in a row and German month names in a
    column.  The function melts this into a tidy DataFrame with one row per
    year-month combination.

    Args:
        filename: Path to the Excel workbook.
        sheet_name: Name of the worksheet to read from.
        top_left_cell: Cell reference (e.g. ``'B29'``) of the top-left corner
            of the table.  This cell's row contains year headers; the rows
            below hold German month names in the same column.

    Returns:
        A DataFrame with columns ``Datum``, ``Jahr``, ``Monat``, ``Anzahl``,
        trimmed to the range between the first and last non-null data points.
    """
    wb = openpyxl.load_workbook(filename, data_only=True)
    try:
        ws = wb[sheet_name]
        col_letter, start_row = coordinate_from_string(top_left_cell)
        start_col = column_index_from_string(col_letter)

        year_columns = _read_year_headers(ws, start_row, start_col)
        records = _read_month_rows(ws, start_row, start_col, year_columns)
    finally:
        wb.close()

    df = pd.DataFrame(records, columns=["Datum", "Jahr", "Monat", "Anzahl"])
    df = df.sort_values("Datum").reset_index(drop=True)
    df = _trim_to_data_range(df)
    df = _keep_only_past_years(df)
    df["Anzahl"] = df["Anzahl"].astype("Int64")
    return df


def _read_year_headers(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    header_row: int,
    start_col: int,
) -> list[tuple[int, int]]:
    """Extract ``(column_index, year)`` pairs from the header row.

    Args:
        ws: The active worksheet.
        header_row: Row number containing year headers.
        start_col: Column number of the month-name column.

    Returns:
        List of tuples mapping column indices to their year values.
    """
    year_columns: list[tuple[int, int]] = []
    for col in range(start_col + 1, ws.max_column + 1):
        value = ws.cell(row=header_row, column=col).value
        if value is not None:
            year_columns.append((col, int(value)))
    return year_columns


def _read_month_rows(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    start_row: int,
    start_col: int,
    year_columns: list[tuple[int, int]],
) -> list[dict[str, object]]:
    """Read 12 month rows and cross-join them with year columns.

    Args:
        ws: The active worksheet.
        start_row: Row of the year headers (months start one row below).
        start_col: Column containing German month names.
        year_columns: Output of :func:`_read_year_headers`.

    Returns:
        List of record dicts ready for DataFrame construction.
    """
    records: list[dict[str, object]] = []
    for i in range(1, 13):
        row_idx = start_row + i
        month_name = ws.cell(row=row_idx, column=start_col).value
        if month_name is None:
            break
        month_name = str(month_name).strip()
        month_number = MONTH_NAME_TO_NUMBER.get(month_name)
        if month_number is None:
            logger.warning("Unknown month name '%s' in row %d, skipping.", month_name, row_idx)
            continue

        for col_idx, year in year_columns:
            value = ws.cell(row=row_idx, column=col_idx).value
            records.append(
                {
                    "Datum": f"{year}-{month_number:02d}",
                    "Jahr": year,
                    "Monat": month_number,
                    "Anzahl": value,
                }
            )
    return records


def _trim_to_data_range(df: pd.DataFrame) -> pd.DataFrame:
    """Remove leading/trailing rows that have no data.

    Keeps NaN rows only when they fall between the first and last non-null
    ``Anzahl`` values, so that gaps within the time series are preserved.

    Args:
        df: Sorted DataFrame with an ``Anzahl`` column.

    Returns:
        Trimmed DataFrame with reset index.
    """
    non_null_mask = df["Anzahl"].notna()
    if not non_null_mask.any():
        return df
    first_valid = non_null_mask.idxmax()
    last_valid = non_null_mask[::-1].idxmax()
    return df.loc[first_valid:last_valid].reset_index(drop=True)


def _keep_only_past_years(df: pd.DataFrame) -> pd.DataFrame:
    """Keep only rows whose year is strictly in the past."""
    current_year = date.today().year
    return df[df["Jahr"] < current_year].reset_index(drop=True)


def main() -> None:
    """Read indicator tables from the AWA Excel workbook and export as CSV."""
    for dataset in DATASETS:
        name = dataset["name"]
        logger.info("Reading %s...", name)
        df = read_monthly_table(
            filename=SOURCE_FILE,
            sheet_name=SHEET_NAME,
            top_left_cell=dataset["top_left_cell"],
        )
        output_path = Path(dataset["output"])
        df.to_csv(output_path, index=False)
        common.update_ftp_and_odsp(str(output_path), "awa/arbeitsinspektorat/", dataset["dataset_id"])
        logger.info("Wrote %d rows to %s.", len(df), output_path)


if __name__ == "__main__":
    logging.basicConfig(level=logging.DEBUG)
    logger.info("Executing %s...", __file__)
    main()
    logger.info("Job successful.")
